# -*- coding: utf-8 -*-
"""
Vascular Health Analyzer - EVA / SUPERNOVA
Versión modernizada con UI premium y gráficos profesionales.

Conserva el 100% de la funcionalidad original:
  - Autenticación (login, registro, recuperación) vía users.py
  - Firma y sello digital por usuario
  - Importación PDF/TXT (incluida captura de curva carótido-femoral)
  - Motor clínico VascularEngine (cálculos, percentiles, fenotipo, riesgo)
  - Generación de PDF profesional de 3 hojas
  - Base JSONL persistente y exportación Excel
  - Panel de administración

Mejoras de esta versión:
  - Tema visual premium (paleta médica, tipografía, cards, botones, badges).
  - Gráfico interactivo Plotly en pantalla con tooltips, zoom y anotaciones.
  - Lámina matplotlib del PDF con calidad editorial (gradientes, paneles
    enriquecidos, fenotipo destacado, comparación de edades, indicador de riesgo).
"""

import datetime
import json
import os
import re
import sys
import traceback
import unicodedata
from io import BytesIO

# ---- Imports con manejo amigable de errores -------------------------------
_MISSING = []
try:
    import matplotlib
    matplotlib.use("Agg")  # backend headless para servidores/Streamlit Cloud
    import matplotlib.pyplot as plt
    from matplotlib.gridspec import GridSpec
    from matplotlib.patches import FancyBboxPatch, Wedge
    from matplotlib.lines import Line2D
    import matplotlib.patheffects as path_effects
except ImportError:
    _MISSING.append("matplotlib")
try:
    import numpy as np
except ImportError:
    _MISSING.append("numpy")
try:
    import pandas as pd
except ImportError:
    _MISSING.append("pandas")
try:
    import openpyxl  # noqa: F401 - requerido para exportación Excel
except ImportError:
    _MISSING.append("openpyxl")
try:
    import streamlit as st
except ImportError:
    print("ERROR: streamlit no esta instalado. Ejecute:\n"
          "  pip install streamlit matplotlib numpy pandas fpdf2 openpyxl pdfplumber Pillow plotly",
          file=sys.stderr)
    sys.exit(1)
try:
    from fpdf import FPDF
except ImportError:
    _MISSING.append("fpdf2")

# Plotly es opcional: si no está, la app cae elegantemente a matplotlib en pantalla.
try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    _PLOTLY_OK = True
except ImportError:
    go = None  # type: ignore
    make_subplots = None  # type: ignore
    _PLOTLY_OK = False

# Modulo local de gestion de usuarios (users.py debe estar al lado de app.py)
UserStore = None
_USERS = None
_USER_ERR = ""
try:
    from users import UserStore as _US
    UserStore = _US
    _USERS = UserStore()
except Exception as _e:
    _USER_ERR = f"{type(_e).__name__}: {_e}"

if _MISSING:
    st.set_page_config(page_title="Faltan dependencias", layout="centered")
    st.error(
        "Faltan paquetes Python para ejecutar la app: **" + ", ".join(_MISSING) + "**.\n\n"
        "Instalelos en su terminal con:\n\n"
        "```\npip install streamlit matplotlib numpy pandas fpdf2 openpyxl pdfplumber Pillow plotly\n```"
    )
    st.stop()

# Lector de PDF (para importar archivos PDF; TXT no requiere librerías adicionales)
try:
    import pdfplumber  # type: ignore
    _PDF_BACKEND = "pdfplumber"
except ImportError:
    try:
        from pypdf import PdfReader  # type: ignore
        _PDF_BACKEND = "pypdf"
    except ImportError:
        try:
            from PyPDF2 import PdfReader  # type: ignore
            _PDF_BACKEND = "pypdf2"
        except ImportError:
            _PDF_BACKEND = None

# Captura opcional de imagen desde PDF (curva carotídeo-femoral)
try:
    import fitz  # PyMuPDF  # type: ignore
    _FITZ_OK = True
except Exception:
    fitz = None  # type: ignore
    _FITZ_OK = False

try:
    from PIL import Image  # type: ignore
    _PIL_OK = True
except Exception:
    Image = None  # type: ignore
    _PIL_OK = False


# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Vascular Health Analyzer - EVA/SUPERNOVA",
    layout="wide",
    page_icon="🫀",
    initial_sidebar_state="expanded",
)

# Paleta clínica (compatibilidad con código original preservada).
COLOR_SUPERNOVA = "#1F6FB2"
COLOR_NORMAL = "#1E8449"
COLOR_EVA = "#C0392B"
COLOR_BG_HEADER = (31, 78, 120)
COLOR_BG_SECTION = (230, 236, 245)
COLOR_BG_ALERT = (253, 237, 236)
COLOR_BG_OK = (232, 245, 233)

# Paleta premium para UI moderna.
THEME = {
    "primary": "#0A2540",       # azul profundo médico
    "primary_2": "#1E3A5F",
    "accent": "#00B4D8",        # cian vibrante
    "accent_2": "#0077B6",
    "supernova": "#1F6FB2",
    "supernova_light": "#D6EAF8",
    "normal": "#1E8449",
    "normal_light": "#D5F5E3",
    "eva": "#C0392B",
    "eva_light": "#FADBD8",
    "warning": "#E67E22",
    "muted": "#5D6D7E",
    "surface": "#FFFFFF",
    "surface_2": "#F8FAFC",
    "border": "#E2E8F0",
    "text": "#0F172A",
    "text_muted": "#64748B",
}


def _inject_premium_css():
    """Inyecta CSS para una experiencia visual profesional y moderna."""
    css = f"""
    <style>
    /* ---- Tipografía global ---- */
    html, body, [class*="css"], .stApp {{
        font-family: 'Inter', 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif !important;
        color: {THEME['text']};
    }}

    /* ---- Fondo general ---- */
    .stApp {{
        background:
            radial-gradient(1100px 600px at 95% -10%, rgba(0,180,216,0.07), transparent 60%),
            radial-gradient(900px 500px at -10% 110%, rgba(10,37,64,0.06), transparent 60%),
            linear-gradient(180deg, #F8FAFC 0%, #EEF2F7 100%);
    }}

    /* ---- Cabecera Streamlit (oculta menú/footer/badge) ---- */
    #MainMenu, footer, header {{visibility: hidden;}}
    .stDeployButton {{display: none;}}

    /* ---- Sidebar ---- */
    section[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, {THEME['primary']} 0%, {THEME['primary_2']} 100%);
    }}
    section[data-testid="stSidebar"] * {{
        color: #E2E8F0 !important;
    }}
    section[data-testid="stSidebar"] .stTextInput input,
    section[data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] > div {{
        background: rgba(255,255,255,0.08) !important;
        color: #FFFFFF !important;
        border: 1px solid rgba(255,255,255,0.18) !important;
        border-radius: 10px !important;
    }}
    section[data-testid="stSidebar"] .stButton button {{
        background: linear-gradient(90deg, {THEME['accent']}, {THEME['accent_2']});
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        padding: 0.55rem 1rem !important;
        box-shadow: 0 4px 14px rgba(0,180,216,0.35);
        transition: transform .15s ease, box-shadow .15s ease;
    }}
    section[data-testid="stSidebar"] .stButton button:hover {{
        transform: translateY(-1px);
        box-shadow: 0 6px 20px rgba(0,180,216,0.45);
    }}
    section[data-testid="stSidebar"] [data-testid="stExpander"] {{
        background: rgba(255,255,255,0.05);
        border: 1px solid rgba(255,255,255,0.10);
        border-radius: 12px;
    }}

    /* ---- Botones principales ---- */
    .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {{
        background: linear-gradient(90deg, {THEME['primary']}, {THEME['accent_2']});
        color: #FFFFFF !important;
        border: none !important;
        border-radius: 12px !important;
        padding: 0.65rem 1.4rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.2px;
        box-shadow: 0 6px 18px rgba(10,37,64,0.18);
        transition: transform .15s ease, box-shadow .15s ease, filter .15s ease;
    }}
    .stButton > button:hover, .stDownloadButton > button:hover, .stFormSubmitButton > button:hover {{
        transform: translateY(-1px);
        box-shadow: 0 10px 24px rgba(10,37,64,0.25);
        filter: brightness(1.05);
    }}

    /* ---- Inputs ---- */
    .stTextInput input, .stNumberInput input, .stTextArea textarea,
    .stSelectbox div[data-baseweb="select"] > div {{
        border-radius: 10px !important;
        border: 1px solid {THEME['border']} !important;
        background: {THEME['surface']} !important;
        transition: border-color .15s ease, box-shadow .15s ease;
    }}
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus {{
        border-color: {THEME['accent']} !important;
        box-shadow: 0 0 0 3px rgba(0,180,216,0.18) !important;
    }}

    /* ---- Tabs ---- */
    .stTabs [data-baseweb="tab-list"] {{
        gap: 6px;
        background: {THEME['surface']};
        padding: 6px;
        border-radius: 12px;
        border: 1px solid {THEME['border']};
    }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 8px !important;
        font-weight: 600 !important;
        color: {THEME['text_muted']} !important;
        padding: 8px 18px !important;
    }}
    .stTabs [aria-selected="true"] {{
        background: linear-gradient(90deg, {THEME['primary']}, {THEME['accent_2']}) !important;
        color: #FFFFFF !important;
    }}

    /* ---- Métricas ---- */
    [data-testid="stMetric"] {{
        background: {THEME['surface']};
        border: 1px solid {THEME['border']};
        border-radius: 14px;
        padding: 14px 16px;
        box-shadow: 0 4px 14px rgba(15,23,42,0.04);
        transition: transform .15s ease, box-shadow .15s ease;
    }}
    [data-testid="stMetric"]:hover {{
        transform: translateY(-2px);
        box-shadow: 0 10px 24px rgba(15,23,42,0.08);
    }}
    [data-testid="stMetricLabel"] {{
        color: {THEME['text_muted']} !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.5px;
        font-size: 11px !important;
    }}
    [data-testid="stMetricValue"] {{
        color: {THEME['primary']} !important;
        font-weight: 800 !important;
    }}

    /* ---- Expanders ---- */
    [data-testid="stExpander"] {{
        background: {THEME['surface']};
        border: 1px solid {THEME['border']} !important;
        border-radius: 14px !important;
        box-shadow: 0 4px 14px rgba(15,23,42,0.04);
    }}
    [data-testid="stExpander"] summary {{
        font-weight: 600 !important;
        color: {THEME['primary']} !important;
    }}

    /* ---- Alertas ---- */
    .stAlert {{
        border-radius: 12px !important;
        border: 1px solid {THEME['border']} !important;
    }}

    /* ---- DataFrame ---- */
    [data-testid="stDataFrame"] {{
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid {THEME['border']};
    }}

    /* ---- Componentes custom ---- */
    .vh-hero {{
        background: linear-gradient(120deg, {THEME['primary']} 0%, {THEME['accent_2']} 60%, {THEME['accent']} 100%);
        border-radius: 18px;
        padding: 26px 30px;
        color: #FFFFFF;
        box-shadow: 0 16px 40px rgba(10,37,64,0.25);
        margin-bottom: 18px;
        position: relative;
        overflow: hidden;
    }}
    .vh-hero::after {{
        content: "🫀";
        position: absolute;
        right: 20px;
        top: 50%;
        transform: translateY(-50%);
        font-size: 80px;
        opacity: 0.14;
    }}
    .vh-hero h1 {{
        font-size: 26px;
        margin: 0 0 6px 0;
        font-weight: 800;
        letter-spacing: 0.2px;
    }}
    .vh-hero p {{
        margin: 0;
        opacity: 0.9;
        font-size: 14px;
    }}

    .vh-card {{
        background: {THEME['surface']};
        border: 1px solid {THEME['border']};
        border-radius: 16px;
        padding: 20px 22px;
        box-shadow: 0 6px 18px rgba(15,23,42,0.05);
        margin-bottom: 14px;
    }}
    .vh-card-title {{
        color: {THEME['primary']};
        font-weight: 700;
        font-size: 15px;
        margin: 0 0 10px 0;
        display: flex;
        align-items: center;
        gap: 8px;
    }}
    .vh-card-title::before {{
        content: "";
        width: 4px;
        height: 18px;
        background: linear-gradient(180deg, {THEME['accent']}, {THEME['primary']});
        border-radius: 4px;
    }}

    .vh-badge {{
        display: inline-block;
        padding: 4px 12px;
        border-radius: 999px;
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 0.4px;
        text-transform: uppercase;
    }}
    .vh-badge-supernova {{
        background: {THEME['supernova_light']}; color: {THEME['supernova']};
    }}
    .vh-badge-normal {{
        background: {THEME['normal_light']}; color: {THEME['normal']};
    }}
    .vh-badge-eva {{
        background: {THEME['eva_light']}; color: {THEME['eva']};
    }}

    .vh-login-wrap {{
        max-width: 520px;
        margin: 30px auto;
    }}
    .vh-login-card {{
        background: rgba(255,255,255,0.95);
        backdrop-filter: blur(8px);
        border: 1px solid {THEME['border']};
        border-radius: 22px;
        padding: 30px 32px;
        box-shadow: 0 24px 60px rgba(10,37,64,0.18);
    }}
    .vh-login-header {{
        text-align: center;
        margin-bottom: 18px;
    }}
    .vh-login-logo {{
        width: 64px;
        height: 64px;
        margin: 0 auto 12px auto;
        border-radius: 18px;
        display: flex;
        align-items: center;
        justify-content: center;
        background: linear-gradient(135deg, {THEME['primary']}, {THEME['accent']});
        color: #FFF;
        font-size: 32px;
        box-shadow: 0 10px 28px rgba(0,180,216,0.30);
    }}
    .vh-login-title {{
        color: {THEME['primary']};
        font-weight: 800;
        font-size: 22px;
        margin: 0;
    }}
    .vh-login-sub {{
        color: {THEME['text_muted']};
        font-size: 13px;
        margin: 4px 0 0 0;
    }}

    .vh-result-pill {{
        display: inline-flex;
        align-items: center;
        gap: 10px;
        padding: 10px 18px;
        border-radius: 14px;
        font-weight: 700;
        font-size: 16px;
        margin-bottom: 14px;
    }}

    /* ---- Encabezados de sección ---- */
    h1, h2, h3 {{
        color: {THEME['primary']};
        font-weight: 700;
    }}

    /* ---- Slim scrollbar ---- */
    ::-webkit-scrollbar {{ width: 10px; height: 10px; }}
    ::-webkit-scrollbar-thumb {{
        background: linear-gradient(180deg, {THEME['accent_2']}, {THEME['primary']});
        border-radius: 10px;
    }}
    </style>
    """
    st.markdown(css, unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# ARCHIVOS POR USUARIO: FIRMA Y SELLO DIGITAL
# ---------------------------------------------------------------------------
APP_DIR = os.path.dirname(os.path.abspath(__file__))
USER_ASSETS_DIR = os.path.join(APP_DIR, "user_assets")
DATA_DIR = os.path.join(APP_DIR, "data")
PATIENT_DB_JSONL = os.path.join(DATA_DIR, "pacientes_vop_registros.jsonl")


def _usuario_seguro(username: str) -> str:
    txt = (username or "usuario").strip().lower()
    txt = re.sub(r"[^a-z0-9_.-]+", "_", txt)
    return txt[:60] or "usuario"


def _guardar_asset_usuario(username: str, uploaded_file, tipo: str):
    """Guarda firma/sello en carpeta aislada por usuario autenticado/registrado."""
    if uploaded_file is None:
        return None
    tipo = "sello" if tipo == "sello" else "firma"
    user_dir = os.path.join(USER_ASSETS_DIR, _usuario_seguro(username))
    os.makedirs(user_dir, exist_ok=True)
    ext = os.path.splitext(getattr(uploaded_file, "name", ""))[1].lower()
    if ext not in (".png", ".jpg", ".jpeg"):
        ext = ".png"
    path = os.path.join(user_dir, f"{tipo}{ext}")
    uploaded_file.seek(0)
    with open(path, "wb") as f:
        f.write(uploaded_file.read())
    return path


def _asset_usuario(username: str, tipo: str):
    """Devuelve la ruta del asset del usuario, si existe. Nunca usa assets de otro usuario."""
    tipo = "sello" if tipo == "sello" else "firma"
    user_dir = os.path.join(USER_ASSETS_DIR, _usuario_seguro(username))
    for ext in (".png", ".jpg", ".jpeg"):
        path = os.path.join(user_dir, f"{tipo}{ext}")
        if os.path.exists(path):
            return path
    return None


def _perfil_usuario_actual(username: str):
    return {
        "firma_path": _asset_usuario(username, "firma"),
        "sello_path": _asset_usuario(username, "sello"),
    }


# ---------------------------------------------------------------------------
# BASE DE DATOS SIMPLE POR USUARIO + EXPORTACIÓN EXCEL
# ---------------------------------------------------------------------------
def _asegurar_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def _guardar_registro_paciente(registro: dict):
    """Guarda un registro clínico en JSONL persistente.

    Cada línea pertenece a un paciente evaluado y queda asociada al usuario
    autenticado. Un médico solo exporta sus registros; el administrador exporta
    todos los usuarios.
    """
    _asegurar_data_dir()
    reg = dict(registro or {})
    reg.setdefault("timestamp", datetime.datetime.now().isoformat(timespec="seconds"))
    with open(PATIENT_DB_JSONL, "a", encoding="utf-8") as f:
        f.write(json.dumps(reg, ensure_ascii=False, default=str) + "\n")


def _cargar_registros_pacientes():
    if not os.path.exists(PATIENT_DB_JSONL):
        return []
    registros = []
    with open(PATIENT_DB_JSONL, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                registros.append(json.loads(line))
            except Exception:
                continue
    return registros


def _df_registros(usuario_actual: str, rol: str):
    registros = _cargar_registros_pacientes()
    if rol != "admin":
        u = _usuario_seguro(usuario_actual)
        registros = [r for r in registros if r.get("usuario_id") == u]
    if not registros:
        return pd.DataFrame()
    df = pd.DataFrame(registros)
    columnas_preferidas = [
        "timestamp", "fecha_estudio", "usuario", "usuario_id", "rol",
        "paciente", "documento", "edad", "sexo", "medico_solicitante",
        "vop_cf_ms", "distancia_cf_cm", "tiempo_transito_cf_ms",
        "pas_mmhg", "pad_mmhg", "pam_mmhg", "pp_mmhg",
        "p10_ms", "p25_ms", "p50_ms", "p75_ms", "p90_ms",
        "edad_vascular", "fenotipo_vascular_unico", "riesgo_cv",
        "lob_vop_mayor_10", "riesgo_elevado_por_vop_mayor_10",
        "fuente_vop", "archivo_importado"
    ]
    columnas = [c for c in columnas_preferidas if c in df.columns] + [c for c in df.columns if c not in columnas_preferidas]
    return df[columnas]


def _excel_bytes_registros(df: pd.DataFrame, nombre_hoja="Pacientes"):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nombre_hoja[:31])
        ws = writer.book[nombre_hoja[:31]]
        # Congelar encabezado y ajustar ancho de columnas para lectura clínica.
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = 10
            letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(max_len + 2, 42)
    return output.getvalue()


# ---------------------------------------------------------------------------
# LÓGICA CLÍNICA (sin cambios funcionales)
# ---------------------------------------------------------------------------
class VascularEngine:
    @staticmethod
    def calcular_vop(distancia_cm: float, tiempo_ms: float) -> float:
        """Calcula VOP cf desde distancia y tiempo, sin factor corrector 0.8.
        Este cálculo queda como respaldo; si el PDF informa VOP cf medida,
        se usa el valor medido por el equipo como dato primario.
        """
        if tiempo_ms <= 0:
            return 0.0
        d_m = distancia_cm / 100
        t_s = tiempo_ms / 1000
        return round(d_m / t_s, 2)

    @staticmethod
    def obtener_percentiles(edad: int, sexo: str):
        base = 5.5 if sexo == "Femenino" else 6.0
        inc = 0.08 * (edad - 20) if edad > 20 else 0
        p50 = base + inc
        return tuple(round(v, 2) for v in (p50 * 0.80, p50 * 0.90, p50, p50 * 1.15, p50 * 1.30))

    @staticmethod
    def clasificar_fenotipo(vop, p10, p90):
        """Clasificacion segun guias ARTERY 2024:
           - SUPERNOVA: VOP < p10 (segun sexo y edad)
           - HVA (Envejecimiento Vascular Saludable): p10 <= VOP <= p90
           - EVA: VOP > p90
        LOB (lesion de organo blanco) es una marca adicional cuando VOP > 10 m/s,
        independiente del fenotipo (se reporta por separado).
        """
        if vop < p10:
            return "SUPERNOVA (Envejecimiento Vascular Supranormal)", COLOR_SUPERNOVA, "Bajo"
        if vop > p90:
            return "EVA (Envejecimiento Vascular Acelerado)", COLOR_EVA, "Alto"
        return "HVA (Envejecimiento Vascular Saludable)", COLOR_NORMAL, "Bajo"

    @staticmethod
    def edad_vascular(vop, sexo):
        base = 5.5 if sexo == "Femenino" else 6.0
        if vop <= base:
            return 20
        return int(round(max(20, min(20 + (vop - base) / 0.08, 100))))

    @staticmethod
    def presion_pulso(pas, pad):
        return round(pas - pad, 1)

    @staticmethod
    def presion_arterial_media(pas, pad):
        return round(pad + (pas - pad) / 3, 1)

    @staticmethod
    def riesgo_cv_global(vop, pas, pad, edad):
        pp = pas - pad
        hta = pas >= 140 or pad >= 90
        if vop > 13 or (vop > 10 and hta and edad >= 60):
            return "Muy Alto"
        if vop > 10 or pp >= 60 or hta:
            return "Alto"
        if vop > 9 or pp >= 50:
            return "Moderado"
        return "Bajo"

    @staticmethod
    def recomendaciones(fenotipo, riesgo, vop, pas, pad):
        recs = []
        if "EVA" in fenotipo:
            recs += [
                "Control intensivo de presion arterial (objetivo < 130/80 mmHg).",
                "Dieta DASH/mediterranea, reduccion de sodio (<5 g/dia).",
                "Actividad fisica aerobica >=150 min/semana + entrenamiento de fuerza.",
                "Perfil lipidico, glicemia y funcion renal en proximas 4 semanas.",
                "Reevaluacion de VOP en 6-12 meses para monitorizar progresion.",
            ]
        elif "SUPERNOVA" in fenotipo:
            recs += [
                "Mantener habitos cardioprotectores actuales.",
                "Reforzar adherencia a actividad fisica regular y dieta equilibrada.",
                "Reevaluacion bianual de rigidez arterial.",
            ]
        else:
            recs += [
                "Promover habitos cardiosaludables (dieta, ejercicio, sueno).",
                "Control anual de presion arterial y perfil metabolico.",
                "Reevaluacion de VOP cada 2 anos o ante nuevos factores de riesgo.",
            ]
        if pas >= 140 or pad >= 90:
            recs.append("Confirmar diagnostico de HTA con MAPA/AMPA e iniciar tratamiento (ESC/ESH).")
        if riesgo in ("Alto", "Muy Alto"):
            recs.append("Derivacion a Cardiologia/Medicina Vascular para evaluacion integral.")
        return recs


# ---------------------------------------------------------------------------
# IMPORTADOR PDF/TXT (lógica clínica intacta del original)
# ---------------------------------------------------------------------------
class GenericReportParser:
    """Extrae datos desde PDF o TXT priorizando VOP carotídeo-femoral.

    Puntos corregidos:
    - No usa VOP radial como referencia.
    - Busca explícitamente VOP cf / carotídeo-femoral / cfPWV.
    - Importa distancia CF, tiempo de tránsito y VOP cf medida desde el archivo original.
    - Si un dato no se detecta, queda en 0/SD para que no aparezcan valores por defecto falsos.
    """

    NEGATIVOS_CF = [
        "radial", "braquial", "brachial", "perifer", "peripheral",
        "aortica", "aórtica", "central", "augmentation", "aix",
        "referencia", "reference", "normal", "umbral", "cutoff", "objetivo",
    ]
    POSITIVOS_CF = [
        "cf", "c-f", "c/f", "carot", "femor", "femoral", "carotido", "carótido",
        "carotideo", "carotídeo", "carotid", "cfpwv", "pwvcf",
    ]

    PATRONES = {
        "nombre": [
            r"(?:Nombre\s+y\s+Apellido|Apellido\s+y\s+Nombre|Paciente|Nombre)\s*[:\-]?\s*([A-Za-zÁÉÍÓÚÑáéíóúñ\.\s]{3,80}?)(?=\s*(?:\n|Edad|DNI|Documento|Sexo|Fecha|$))",
        ],
        "documento": [
            r"(?:DNI|Documento|D\.N\.I\.?|N[°º]\s*Documento|CI|Cedula|Cédula)\s*[:\-]?\s*([\d\.\-]{6,15})",
        ],
        "edad": [
            r"Edad\s*[:\-]?\s*(\d{1,3})\s*(?:a[nñ]os|años|a\.?)?",
        ],
        "sexo": [
            r"Sexo\s*[:\-]?\s*(Masculino|Femenino|M|F|Mas|Fem)\b",
        ],
    }

    @staticmethod
    def _norm_num(v):
        if v is None:
            return None
        try:
            txt = str(v).strip().replace(" ", "")
            txt = txt.replace("m/s", "").replace("ms", "").replace("cm", "")
            if "," in txt and "." in txt:
                txt = txt.replace(".", "").replace(",", ".")
            else:
                txt = txt.replace(",", ".")
            return float(txt)
        except Exception:
            return None

    @staticmethod
    def _ventana(texto, ini, fin, n=120):
        return texto[max(0, ini-n):min(len(texto), fin+n)].lower()

    @classmethod
    def _es_cf(cls, ventana: str) -> bool:
        if any(x in ventana for x in cls.NEGATIVOS_CF):
            if any(x in ventana for x in ["radial", "braquial", "brachial", "perifer", "peripheral", "central", "aortica", "aórtica"]):
                return False
            if any(x in ventana for x in ["referencia", "reference", "normal", "umbral", "cutoff", "objetivo"]):
                return False
        return any(x in ventana for x in cls.POSITIVOS_CF)

    @staticmethod
    def _rango(v, lo, hi):
        return v is not None and lo <= v <= hi

    @staticmethod
    def _lineas_reales(txt: str):
        """Devuelve líneas no vacías conservando el orden original del archivo."""
        if not txt:
            return []
        return [re.sub(r"\s+", " ", x).strip() for x in str(txt).splitlines() if x and x.strip()]

    @staticmethod
    def _normalizar_ocr(txt: str) -> str:
        """Normaliza variantes frecuentes de PDF/TXT sin formato y errores de OCR."""
        if txt is None:
            return ""
        txt = str(txt)
        txt = txt.replace(" ", " ").replace("│", "|").replace("¦", "|")
        txt = txt.replace("Carótida", "Carotida").replace("Carótido", "Carotido")
        txt = txt.replace("carótida", "carotida").replace("carótido", "carotido")
        txt = txt.replace("Fem.", "Fem.")
        return txt

    @staticmethod
    def _limpiar_texto(txt: str) -> str:
        txt = txt.replace(" ", " ").replace("\t", " ")
        txt = txt.replace("：", ":").replace("＝", "=").replace("–", "-").replace("—", "-")
        txt = re.sub(r"[ ]+", " ", txt)
        txt = re.sub(r"\n{3,}", "\n\n", txt)
        return txt

    @staticmethod
    def _nombre_desde_archivo(file_obj):
        """Respaldo cuando el TXT/PDF no trae etiqueta clara de paciente."""
        nombre_archivo = getattr(file_obj, "name", "") or ""
        base = re.sub(r"\.[A-Za-z0-9]{1,6}$", "", nombre_archivo).strip()
        base = re.sub(r"[_-]+", " ", base)
        base = re.sub(r"\b(?:PDF|TXT|INFORME|REPORTE|VOP|EVA|SUPERNOVA)\b", " ", base, flags=re.IGNORECASE)
        base = re.sub(r"\b(?:copia|copy|final|corregido|corr|nuevo|curva|cf)\b", " ", base, flags=re.IGNORECASE)
        base = re.sub(r"\b\d+\b", " ", base)
        base = re.sub(r"\s+", " ", base).strip(" -_:;,.()[]{}")
        if len(base) >= 5 and re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]", base):
            return base.upper() if base.isupper() else base.title()
        return None

    @classmethod
    def _extraer_nombre_paciente(cls, texto, file_obj=None):
        """Extractor tolerante para nombre/apellido del paciente."""
        t = cls._limpiar_texto(texto or "")
        patrones = [
            r"(?:Apellido\s*y\s*Nombre|Nombre\s*y\s*Apellido|Paciente|Nombre\s*Completo|ApyNom|Ape\.?\s*y\s*Nom\.?)\s*(?:[:=\-]|\|)?\s*([^\n\r|]{3,90})",
            r"(?:Patient|Name)\s*(?:[:=\-]|\|)?\s*([^\n\r|]{3,90})",
        ]
        cortes = r"\b(?:Edad|DNI|Documento|Sexo|Fecha|Obra\s*Social|M[eé]dico|Diagn[oó]stico|PAS|PAD|VOP|Vel\.|TCF|Dist\.)\b"
        for pat in patrones:
            for m in re.finditer(pat, t, flags=re.IGNORECASE):
                cand = m.group(1).strip(" :-|,.;")
                cand = re.split(cortes, cand, maxsplit=1, flags=re.IGNORECASE)[0].strip(" :-|,.;")
                cand = re.sub(r"\s+", " ", cand)
                if 3 <= len(cand) <= 90 and re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]", cand):
                    if not re.search(r"^(edad|sexo|dni|documento|fecha|vel|vop|tcf|dist)", cand, flags=re.IGNORECASE):
                        return cand
        lineas = [x.strip(" :-|,.;") for x in t.splitlines() if x.strip(" :-|,.;")]
        for i, lin in enumerate(lineas[:-1]):
            if re.fullmatch(r"(?:Paciente|Nombre|Apellido\s*y\s*Nombre|Nombre\s*y\s*Apellido|Nombre\s*Completo)", lin, flags=re.IGNORECASE):
                cand = re.sub(r"\s+", " ", lineas[i+1]).strip(" :-|,.;")
                if 3 <= len(cand) <= 90 and re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]", cand):
                    return cand
        if file_obj is not None:
            return cls._nombre_desde_archivo(file_obj)
        return None

    @staticmethod
    def extraer_texto(file_obj) -> str:
        file_obj.seek(0)
        nombre = getattr(file_obj, "name", "") or ""
        es_txt = nombre.lower().endswith(".txt")
        if es_txt:
            raw = file_obj.read()
            if isinstance(raw, bytes):
                for enc in ("utf-8", "latin-1", "cp1252"):
                    try:
                        return GenericReportParser._limpiar_texto(raw.decode(enc, errors="ignore"))
                    except Exception:
                        continue
                return GenericReportParser._limpiar_texto(raw.decode("utf-8", errors="ignore"))
            return GenericReportParser._limpiar_texto(str(raw))

        if _PDF_BACKEND is None:
            raise RuntimeError("Instale 'pdfplumber' o 'pypdf' para leer PDFs. Los archivos TXT no requieren librerías adicionales.")
        file_obj.seek(0)
        txt = ""
        if _PDF_BACKEND == "pdfplumber":
            with pdfplumber.open(file_obj) as pdf:
                for p in pdf.pages:
                    txt += (p.extract_text(x_tolerance=2, y_tolerance=3) or "") + "\n"
                    try:
                        for table in p.extract_tables() or []:
                            for row in table or []:
                                vals = [str(c).strip() for c in row if c not in (None, "")]
                                if vals:
                                    txt += " | ".join(vals) + "\n"
                    except Exception:
                        pass
        else:
            r = PdfReader(file_obj)
            for p in r.pages:
                txt += (p.extract_text() or "") + "\n"
        return GenericReportParser._limpiar_texto(txt)

    @classmethod
    def _extraer_pa(cls, texto):
        patrones_par = [
            r"(?:TA|PA|Presi[oó]n\s+Arterial|Brachial\s+BP|BP)\s*[:\-=]?\s*(\d{2,3})\s*[/\\]\s*(\d{2,3})\s*(?:mmHg)?",
            r"(?:Sist[oó]lica|PAS|SBP)[^\d]{0,20}(\d{2,3}).{0,60}(?:Diast[oó]lica|PAD|DBP)[^\d]{0,20}(\d{2,3})",
            r"(\d{2,3})\s*[/\\]\s*(\d{2,3})\s*mmHg",
        ]
        for pat in patrones_par:
            for m in re.finditer(pat, texto, flags=re.IGNORECASE | re.DOTALL):
                pas, pad = int(m.group(1)), int(m.group(2))
                ven = cls._ventana(texto, m.start(), m.end(), 100)
                if any(x in ven for x in ["referencia", "normal", "objetivo", "umbral", "percentil"]):
                    continue
                if 70 <= pas <= 260 and 30 <= pad <= 160 and pas > pad:
                    return pas, pad
        pas = None; pad = None
        for pat in [r"(?:PAS|TAS|SBP|Sist[oó]lica)\s*[:\-=]?\s*(\d{2,3})"]:
            m = re.search(pat, texto, flags=re.IGNORECASE)
            if m: pas = int(m.group(1))
        for pat in [r"(?:PAD|TAD|DBP|Diast[oó]lica)\s*[:\-=]?\s*(\d{2,3})"]:
            m = re.search(pat, texto, flags=re.IGNORECASE)
            if m: pad = int(m.group(1))
        if pas and pad and 70 <= pas <= 260 and 30 <= pad <= 160 and pas > pad:
            return pas, pad
        return None, None

    @classmethod
    def _buscar_numero_cf(cls, texto, tipo):
        """tipo: distancia, tiempo o vop. Devuelve el mejor valor CF, nunca radial."""
        texto = cls._limpiar_texto(texto or "")

        if tipo == "distancia":
            exactos = [
                r"(?is)\bD\s*i\s*s\s*t\s*\.?\s*(?:\||:|;|,|-|\s)*C\s*a\s*r\s*\.?\s*(?:\||:|;|,|-|\s)*F\s*e\s*m\s*\.?\s*(?:=|:|-|\||\s)*([0-9]{2,3}(?:[,.][0-9]+)?)\s*(cm|m)?",
                r"(?is)\bDist\.?\s*Car\.?\s*Fem\.?\s*(?:=|:|-|\||\s)*([0-9]{2,3}(?:[,.][0-9]+)?)\s*(cm|m)?",
                r"(?is)\bDist\.?\s*Car\.?\s*Fem\.?[^0-9\n]{0,60}([0-9]{2,3}(?:[,.][0-9]+)?)\s*(cm|m)?",
                r"(?is)\bDistancia\s*Car[oó]tid[ao]\s*[-/]?\s*Femoral\b[^0-9\n]{0,60}([0-9]{2,3}(?:[,.][0-9]+)?)\s*(cm|m)?",
                r"(?is)\bDistancia\s*Car[oó]tido\s*[-/]?\s*Femoral\s*REAL\b[^0-9\n]{0,60}([0-9]{2,3}(?:[,.][0-9]+)?)\s*(cm|m)?",
            ]
            lo, hi = 20, 200
        elif tipo == "tiempo":
            exactos = [
                r"\bTCF\b\s*(?:=|:|-)?\s*([\d.,]+)\s*(mseg|ms|msec|seg|s)?",
                r"\bTiempo\s*de\s*Tr[aá]nsito\s*CF\b\s*(?:=|:|-)?\s*([\d.,]+)\s*(mseg|ms|msec|seg|s)?",
            ]
            lo, hi = 10, 300
        else:
            exactos = [
                r"(?:Vel\.?\s*Car[oó]tida\s*[-/]?\s*Femoral|Vel\.?\s*Car[oó]tido\s*[-/]?\s*Femoral|Velocidad\s*Car[oó]tida\s*[-/]?\s*Femoral)[^\d\n|]{0,40}([\d.,]+)\s*(?:m\s*/\s*s|m/s)?",
                r"(?:VOP|PWV)\s*(?:Car[oó]tida|Car[oó]tido|Carotid)\s*[-/]?\s*Femoral[^\d\n|]{0,40}([\d.,]+)\s*(?:m\s*/\s*s|m/s)?",
            ]
            lo, hi = 3, 25

        for pat in exactos:
            for m in re.finditer(pat, texto, flags=re.IGNORECASE | re.DOTALL):
                val = cls._norm_num(m.group(1))
                if val is None:
                    continue
                unidad = ""
                if tipo in ("distancia", "tiempo") and m.lastindex and m.lastindex >= 2 and m.group(2):
                    unidad = m.group(2).lower()
                if tipo == "distancia" and unidad == "m":
                    val = val * 100
                if tipo == "tiempo" and unidad in ("s", "seg"):
                    val = val * 1000
                if lo <= val <= hi:
                    return round(val, 2)

        if tipo == "distancia":
            patrones = [
                r"(?:distancia\s*car[oó]tido\s*[-/]?\s*femoral|distance|dcf|d\s*cf|c\s*[-/]?f\s*distance|carotid\s*[-/]?femoral\s*distance|car[oó]tido\s*[-/]?femoral\s*distancia)[^\d\n|]{0,40}([\d.,]+)\s*(cm|m)?",
                r"(?:carotid|car[oó]tido|car[oó]tida).{0,50}(?:femoral).{0,50}([\d.,]+)\s*(cm|m)?",
            ]
            lo, hi = 20, 200
        elif tipo == "tiempo":
            patrones = [
                r"(?:tiempo\s*(?:de)?\s*tr[aá]nsito\s*(?:car[oó]tido\s*[-/]?\s*femoral)?|transit\s*time|pulse\s*transit\s*time|ptt|tt|delta\s*t|[Δ∆]\s*t)[^\d\n|]{0,40}([\d.,]+)\s*(ms|mseg|msec|s|seg)?",
                r"(?:carotid|car[oó]tido|car[oó]tida).{0,80}(?:femoral).{0,80}(?:tiempo|transit|tt|ptt).{0,40}([\d.,]+)\s*(ms|mseg|msec|s|seg)?",
            ]
            lo, hi = 10, 300
        else:
            patrones = [
                r"(?:cf\s*[-/]?\s*pwv|cfpwv|pwvcf|pwv\s*cf|vop\s*cf|vop\s*c\s*[-/]?\s*f)[^\d\n|]{0,40}([\d.,]+)\s*(?:m\s*/\s*s|m/s)?",
                r"(?:velocidad\s+de\s+onda\s+de\s+pulso|pulse\s+wave\s+velocity|PWV|VOP).{0,80}(?:car[oó]tido|car[oó]tida|carotid|femoral|cf).{0,40}([\d.,]+)\s*(?:m\s*/\s*s|m/s)?",
                r"(?:car[oó]tido|car[oó]tida|carotid).{0,40}(?:femoral).{0,80}(?:VOP|PWV|velocidad|vel\.)[^\d\n|]{0,40}([\d.,]+)\s*(?:m\s*/\s*s|m/s)?",
            ]
            lo, hi = 3, 25

        candidatos = []
        for pat in patrones:
            for m in re.finditer(pat, texto, flags=re.IGNORECASE | re.DOTALL):
                val = cls._norm_num(m.group(1))
                if val is None:
                    continue
                unidad = m.group(2).lower() if (tipo in ("distancia", "tiempo") and m.lastindex and m.lastindex >= 2 and m.group(2)) else ""
                if tipo == "distancia" and unidad == "m":
                    val = val * 100
                if tipo == "tiempo" and unidad in ("s", "seg"):
                    val = val * 1000
                ven = cls._ventana(texto, m.start(), m.end(), 160)
                if tipo in ("distancia", "tiempo"):
                    if any(x in ven for x in ["radial", "braquial", "brachial", "perifer", "peripheral"]):
                        continue
                    score = 2 if cls._es_cf(ven) else 1
                else:
                    if not cls._es_cf(ven):
                        continue
                    score = 3
                if lo <= val <= hi:
                    candidatos.append((score, m.start(), round(val, 2)))
        if not candidatos:
            if tipo == "distancia":
                for lin in cls._lineas_reales(texto):
                    lin_norm = cls._normalizar_ocr(lin)
                    if re.search(r"(?i)dist\.?\s*car\.?\s*fem\.?", lin_norm) or (re.search(r"(?i)dist", lin_norm) and re.search(r"(?i)car", lin_norm) and re.search(r"(?i)fem", lin_norm)):
                        nums = re.findall(r"(?<![A-Za-z])([0-9]{2,3}(?:[,.][0-9]+)?)(?![A-Za-z])", lin_norm)
                        for num in nums:
                            val = cls._norm_num(num)
                            if val is not None and 20 <= val <= 200:
                                return round(val, 2)
            return None
        candidatos.sort(key=lambda x: (-x[0], x[1]))
        return candidatos[0][2]

    @staticmethod
    def _pixmap_to_png_bytes(page, crop, zoom=2.8):
        """Renderiza una región del PDF a PNG bytes."""
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=crop, alpha=False)
        return pix.tobytes("png")

    @staticmethod
    def _render_page_to_pil(page, zoom=2.5):
        """Renderiza página completa a PIL.Image para localizar la curva cuando el PDF es imagen."""
        if not _PIL_OK:
            return None
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        return img

    @classmethod
    def _capturar_por_pixeles_curva_cf(cls, page):
        """Localiza el panel CF por colores de las curvas roja/verde cuando el texto no es seleccionable."""
        if not _PIL_OK:
            return None
        try:
            img = cls._render_page_to_pil(page, zoom=2.8)
            if img is None:
                return None
            import numpy as _np
            arr = _np.asarray(img)
            r = arr[:, :, 0].astype(int)
            g = arr[:, :, 1].astype(int)
            b = arr[:, :, 2].astype(int)
            mask_red = (r > 135) & (g < 120) & (b < 120) & ((r - g) > 45)
            mask_green = (g > 95) & (r < 140) & (b < 140) & ((g - r) > 25)
            mask = mask_red | mask_green
            ys, xs = _np.where(mask)
            if len(xs) < 80:
                return None
            h, w = mask.shape
            block = max(20, min(h, w) // 30)
            bx = xs // block
            by = ys // block
            keys, counts = _np.unique(by * 10000 + bx, return_counts=True)
            key = int(keys[int(_np.argmax(counts))])
            cy = (key // 10000) * block + block // 2
            cx = (key % 10000) * block + block // 2
            radio_x = max(260, int(w * 0.35))
            radio_y = max(180, int(h * 0.25))
            near = (abs(xs - cx) < radio_x) & (abs(ys - cy) < radio_y)
            if near.sum() < 50:
                near = _np.ones_like(xs, dtype=bool)
            x0, x1 = xs[near].min(), xs[near].max()
            y0, y1 = ys[near].min(), ys[near].max()
            pad_left = int(0.36 * (x1 - x0 + 1)) + 90
            pad_right = int(1.15 * (x1 - x0 + 1)) + 180
            pad_top = int(0.30 * (y1 - y0 + 1)) + 70
            pad_bottom = int(0.60 * (y1 - y0 + 1)) + 110
            crop = (
                max(0, int(x0 - pad_left)),
                max(0, int(y0 - pad_top)),
                min(w, int(x1 + pad_right)),
                min(h, int(y1 + pad_bottom)),
            )
            max_w = int(w * 0.65)
            max_h = int(h * 0.55)
            cx0, cy0, cx1, cy1 = crop
            if (cx1 - cx0) > max_w:
                cx1 = min(w, cx0 + max_w)
            if (cy1 - cy0) > max_h:
                cy1 = min(h, cy0 + max_h)
            crop_img = img.crop((cx0, cy0, cx1, cy1))
            buf = BytesIO()
            crop_img.save(buf, format="PNG")
            return buf.getvalue()
        except Exception:
            return None

    @staticmethod
    def _score_imagen_cf_pil(img):
        """Puntúa si una imagen/crop parece contener el panel de curva CF."""
        if not _PIL_OK or img is None:
            return 0
        try:
            import numpy as _np
            im = img.convert("RGB")
            if max(im.size) > 1200:
                im.thumbnail((1200, 1200))
            arr = _np.asarray(im)
            if arr.size == 0:
                return 0
            r = arr[:, :, 0].astype(int)
            g = arr[:, :, 1].astype(int)
            b = arr[:, :, 2].astype(int)
            mask_red = (r > 130) & (g < 135) & (b < 135) & ((r - g) > 35)
            mask_green = (g > 85) & (r < 155) & (b < 155) & ((g - r) > 18)
            red_n = int(mask_red.sum())
            green_n = int(mask_green.sum())
            h, w = arr.shape[:2]
            area = max(1, h * w)
            color_density = (red_n + green_n) / area
            score = 0
            if red_n > 30 and green_n > 30:
                score += 100
            score += min(80, int(color_density * 20000))
            if 1.0 <= (w / max(1, h)) <= 2.4:
                score += 20
            if w >= 180 and h >= 120:
                score += 20
            return score
        except Exception:
            return 0

    @classmethod
    def _recortar_panel_desde_imagen_pil(cls, img):
        """Recorta el panel CF dentro de una imagen de página o imagen embebida."""
        if not _PIL_OK or img is None:
            return None
        try:
            import numpy as _np
            im = img.convert("RGB")
            arr = _np.asarray(im)
            r = arr[:, :, 0].astype(int)
            g = arr[:, :, 1].astype(int)
            b = arr[:, :, 2].astype(int)
            mask_red = (r > 130) & (g < 135) & (b < 135) & ((r - g) > 35)
            mask_green = (g > 85) & (r < 155) & (b < 155) & ((g - r) > 18)
            mask = mask_red | mask_green
            ys, xs = _np.where(mask)
            if len(xs) < 60:
                return im
            x0, x1 = int(xs.min()), int(xs.max())
            y0, y1 = int(ys.min()), int(ys.max())
            w, h = im.size
            dx = x1 - x0 + 1
            dy = y1 - y0 + 1
            cx0 = max(0, x0 - int(dx * 0.45) - 120)
            cy0 = max(0, y0 - int(dy * 0.55) - 80)
            cx1 = min(w, x1 + int(dx * 1.10) + 220)
            cy1 = min(h, y1 + int(dy * 0.85) + 140)
            crop = im.crop((cx0, cy0, cx1, cy1))
            return crop
        except Exception:
            return img

    @classmethod
    def _capturar_imagen_embebida_cf(cls, doc):
        """Busca imágenes embebidas dentro del PDF y devuelve el panel CF."""
        if not (_FITZ_OK and _PIL_OK):
            return None
        try:
            best_img = None
            best_score = 0
            for page in doc:
                for info in page.get_images(full=True) or []:
                    xref = info[0]
                    try:
                        base = doc.extract_image(xref)
                        data = base.get("image")
                        if not data:
                            continue
                        img = Image.open(BytesIO(data)).convert("RGB")
                        score = cls._score_imagen_cf_pil(img)
                        if score > best_score:
                            best_score = score
                            best_img = img.copy()
                    except Exception:
                        continue
            if best_img is not None and best_score >= 100:
                panel = cls._recortar_panel_desde_imagen_pil(best_img)
                buf = BytesIO()
                panel.save(buf, format="PNG")
                return buf.getvalue()
        except Exception:
            return None
        return None

    @classmethod
    def _capturar_pagina_pdf_respaldo(cls, doc):
        """Respaldo seguro: si no se localiza el panel CF, captura una página completa."""
        if doc is None or len(doc) == 0 or not _FITZ_OK:
            return None
        try:
            idx = 0
            claves = (
                "carot", "femor", "cfpwv", "vop", "pwv",
                "dist. car", "dist car", "tcf", "vel. car"
            )
            for i, page in enumerate(doc):
                try:
                    txt = (page.get_text("text") or "").lower()
                    txt_norm = (txt.replace("ó", "o").replace("í", "i")
                                  .replace("á", "a").replace("é", "e").replace("ú", "u"))
                    if any(k in txt_norm for k in claves):
                        idx = i
                        break
                except Exception:
                    continue
            page = doc[idx]
            pr = page.rect
            return cls._pixmap_to_png_bytes(page, fitz.Rect(pr.x0, pr.y0, pr.x1, pr.y1), zoom=2.1)
        except Exception:
            return None

    @classmethod
    def capturar_curva_cf_pdf(cls, file_obj):
        """Devuelve PNG bytes del sector gráfico VOP carótido-femoral del PDF."""
        nombre = getattr(file_obj, "name", "") or ""
        if not nombre.lower().endswith(".pdf") or not _FITZ_OK:
            return None
        doc = None
        try:
            file_obj.seek(0)
            data = file_obj.read()
            doc = fitz.open(stream=data, filetype="pdf")
            claves = [
                "Vel. Carótida-Femoral", "Vel. Carotida-Femoral",
                "Vel Carotida Femoral", "Velocidad Carotida Femoral",
                "Velocidad Carótida Femoral", "carotida-femoral", "carótida-femoral",
                "carotido-femoral", "carótido-femoral", "VOP CF", "cfPWV",
                "TCF", "T C F", "Dist. Car. Fem.", "Dist Car Fem",
                "Dist. Car Fem", "Dist Car. Fem", "Car. Fem"
            ]
            claves_norm = [
                k.lower().replace("ó", "o").replace("í", "i").replace("á", "a")
                 .replace("é", "e").replace("ú", "u") for k in claves
            ]

            for page in doc:
                rects = []
                for key in claves:
                    try:
                        for rr in page.search_for(key):
                            rects.append(fitz.Rect(rr))
                    except Exception:
                        pass
                if not rects:
                    for b in (page.get_text("blocks") or []):
                        txt = str(b[4]).lower()
                        norm = (txt.replace("ó", "o").replace("í", "i")
                                   .replace("á", "a").replace("é", "e").replace("ú", "u"))
                        if any(k in norm for k in claves_norm):
                            rects.append(fitz.Rect(b[:4]))
                if rects:
                    r = rects[0]
                    for rr in rects[1:]:
                        r |= rr
                    pr = page.rect
                    crop = fitz.Rect(
                        max(pr.x0, r.x0 - max(145, pr.width * 0.18)),
                        max(pr.y0, r.y0 - max(145, pr.height * 0.16)),
                        min(pr.x1, r.x1 + max(520, pr.width * 0.48)),
                        min(pr.y1, r.y1 + max(300, pr.height * 0.30)),
                    )
                    if crop.width < pr.width * 0.52:
                        crop.x1 = min(pr.x1, crop.x0 + pr.width * 0.64)
                    if crop.height < pr.height * 0.32:
                        crop.y1 = min(pr.y1, crop.y0 + pr.height * 0.42)
                    return cls._pixmap_to_png_bytes(page, crop, zoom=3.4)

            png_emb = cls._capturar_imagen_embebida_cf(doc)
            if png_emb:
                return png_emb

            for page in doc:
                png = cls._capturar_por_pixeles_curva_cf(page)
                if png:
                    return png

            respaldo = cls._capturar_pagina_pdf_respaldo(doc)
            if respaldo:
                return respaldo
            return None
        except Exception:
            return None
        finally:
            try:
                if doc is not None:
                    doc.close()
            except Exception:
                pass

    @classmethod
    def parsear(cls, file_obj) -> dict:
        texto = cls.extraer_texto(file_obj)
        curva_png = cls.capturar_curva_cf_pdf(file_obj)
        out = {"_texto_crudo": texto, "_curva_cf_png": curva_png}

        out["nombre"] = cls._extraer_nombre_paciente(texto, file_obj)

        for campo in ("documento", "edad", "sexo"):
            v = None
            for pat in cls.PATRONES[campo]:
                m = re.search(pat, texto, flags=re.IGNORECASE)
                if m:
                    v = m.group(1).strip()
                    break
            out[campo] = v

        if out.get("sexo"):
            out["sexo"] = "Femenino" if str(out["sexo"]).upper().startswith("F") else "Masculino"
        if out.get("edad"):
            try:
                out["edad"] = int(out["edad"])
            except ValueError:
                out["edad"] = None

        out["pas"], out["pad"] = cls._extraer_pa(texto)
        out["distancia"] = cls._buscar_numero_cf(texto, "distancia")
        out["tiempo"] = cls._buscar_numero_cf(texto, "tiempo")
        out["vop"] = cls._buscar_numero_cf(texto, "vop")

        if out.get("vop") is None and out.get("distancia") and out.get("tiempo"):
            out["vop_recalculada"] = VascularEngine.calcular_vop(out["distancia"], out["tiempo"])
        else:
            out["vop_recalculada"] = None

        if out.get("nombre"):
            out["nombre"] = re.sub(r"\s+", " ", out["nombre"]).strip(" :-")

        if out.get("pas") and out.get("pad") and out["pad"] >= out["pas"]:
            out["pas"], out["pad"] = None, None
        return out


# ---------------------------------------------------------------------------
# UTILIDADES
# ---------------------------------------------------------------------------
def safe_latin1(t):
    if t is None:
        return ""
    rep = {"“": '"', "”": '"', "‘": "'", "’": "'", "–": "-", "—": "-",
           "•": "-", "…": "...", "≥": ">=", "≤": "<=", "±": "+/-"}
    for k, v in rep.items():
        t = t.replace(k, v)
    t = unicodedata.normalize("NFKD", t)
    return t.encode("latin-1", "ignore").decode("latin-1")


# ---------------------------------------------------------------------------
# GRÁFICA DIDÁCTICA PROFESIONAL (matplotlib calidad editorial - PDF)
# ---------------------------------------------------------------------------
def construir_grafico_didactico(edad, sexo, vop, p10, p25, p50, p75, p90,
                                color_paciente, edad_vasc,
                                pas=None, pad=None, riesgo=None):
    """Lámina didáctica editorial para el PDF.

    Estructura: 6 paneles distribuidos en grid 3x3:
      [0,0-1] curva de rigidez por edad/sexo con bandas y marcador del paciente
      [0,2]   panel fenotipo (card grande con badge y rangos)
      [1,0]   barras horizontales de percentiles
      [1,1]   comparación edad cronológica vs vascular
      [1,2]   gauge de riesgo cardiovascular
      [2,:]   franja informativa con presión arterial y métricas clave

    Mantiene la firma de la versión anterior (acepta nuevos parámetros opcionales).
    """
    plt.rcParams.update({
        "font.family": "DejaVu Sans",
        "axes.edgecolor": "#475569",
        "axes.titlesize": 10.5,
        "axes.titleweight": "bold",
        "axes.labelsize": 8.8,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "xtick.color": "#475569",
        "ytick.color": "#475569",
        "xtick.labelsize": 8,
        "ytick.labelsize": 8,
    })

    fig = plt.figure(figsize=(11.6, 9.2), dpi=180, facecolor="white")
    gs = GridSpec(3, 3,
                  width_ratios=[1.3, 1.0, 1.0],
                  height_ratios=[1.20, 0.95, 0.42],
                  wspace=0.34, hspace=0.55)

    # ============== Panel 1: curva de rigidez arterial ==============
    ax = fig.add_subplot(gs[0, 0:2])
    edades = np.linspace(20, 80, 250)
    base = 5.5 if sexo == "Femenino" else 6.0
    cp50 = base + 0.08 * (edades - 20)
    cp10 = cp50 * 0.80
    cp25 = cp50 * 0.90
    cp75 = cp50 * 1.15
    cp90 = cp50 * 1.30

    # Bandas con degradado visual (alpha distinto por banda)
    ax.fill_between(edades, 0, cp10, color=COLOR_SUPERNOVA, alpha=0.18, label="SUPERNOVA (< p10)")
    ax.fill_between(edades, cp10, cp25, color="#5DADE2", alpha=0.12)
    ax.fill_between(edades, cp25, cp75, color=COLOR_NORMAL, alpha=0.15, label="HVA (p10-p90)")
    ax.fill_between(edades, cp75, cp90, color="#E59866", alpha=0.14)
    ax.fill_between(edades, cp90, cp90 * 1.6, color=COLOR_EVA, alpha=0.20, label="EVA (> p90)")

    # Línea umbral 10 m/s (riesgo)
    ax.axhline(10, color="#8B0000", lw=1.4, ls="--", alpha=0.78)
    ax.text(21, 10.25, "VOP > 10 m/s = riesgo CV elevado por rigidez arterial",
            fontsize=7.2, color="#8B0000", fontweight="bold")

    # Curvas de percentiles
    ax.plot(edades, cp50, color="#1F2937", lw=2.2, label="p50")
    ax.plot(edades, cp25, color="#334155", lw=1.0, ls=":", alpha=0.75, label="p25 / p75")
    ax.plot(edades, cp75, color="#334155", lw=1.0, ls=":", alpha=0.75)
    ax.plot(edades, cp10, color=COLOR_SUPERNOVA, lw=1.3, ls="--", alpha=0.9)
    ax.plot(edades, cp90, color=COLOR_EVA, lw=1.3, ls="--", alpha=0.9)

    # Marcador del paciente con halo
    halo = ax.scatter(edad, vop, s=380, color=color_paciente, alpha=0.18, zorder=5)
    sc = ax.scatter(edad, vop, color=color_paciente, s=150, zorder=7,
                    edgecolor="white", linewidth=2.2, label="Paciente")
    sc.set_path_effects([path_effects.withStroke(linewidth=3, foreground="white")])

    # Anotación elegante
    off_x = 5 if edad < 65 else -22
    ax.annotate(f"Paciente\n{edad} a / {vop} m/s",
                xy=(edad, vop), xytext=(edad + off_x, vop + 1.4),
                fontsize=8, fontweight="bold", color=color_paciente,
                bbox=dict(boxstyle="round,pad=0.4", fc="white",
                          ec=color_paciente, lw=1.1, alpha=0.95),
                arrowprops=dict(arrowstyle="-|>", color=color_paciente, lw=1.2,
                                connectionstyle="arc3,rad=0.18"))

    ax.set_title(f"Curva de rigidez arterial — Sexo {sexo}", pad=10, color="#0F172A")
    ax.set_xlabel("Edad cronológica (años)")
    ax.set_ylabel("VOP carotídeo-femoral (m/s)")
    ax.set_xlim(20, 82)
    ax.set_ylim(2.5, max(16, vop + 3.5))
    ax.grid(True, ls=":", alpha=0.35, color="#94A3B8")
    leg = ax.legend(loc="upper left", fontsize=7, framealpha=0.94,
                    fancybox=True, edgecolor="#CBD5E1")
    leg.get_frame().set_linewidth(0.6)

    # ============== Panel 2: card fenotipo vascular ==============
    ax3 = fig.add_subplot(gs[0, 2])
    if vop < p10:
        fenotipo_graf = "SUPERNOVA"; c_fen = COLOR_SUPERNOVA
        rango_txt = "VOP < p10"
        sub = "Envejecimiento\nVascular Supranormal"
    elif vop > p90:
        fenotipo_graf = "EVA"; c_fen = COLOR_EVA
        rango_txt = "VOP > p90"
        sub = "Envejecimiento\nVascular Acelerado"
    else:
        fenotipo_graf = "HVA"; c_fen = COLOR_NORMAL
        rango_txt = "p10 <= VOP <= p90"
        sub = "Envejecimiento\nVascular Saludable"

    ax3.axis("off")
    # Tarjeta de fondo
    card = FancyBboxPatch((0.02, 0.05), 0.96, 0.90,
                          boxstyle="round,pad=0.02,rounding_size=0.04",
                          fc="white", ec=c_fen, lw=2.2,
                          transform=ax3.transAxes, zorder=1)
    ax3.add_patch(card)
    # Franja superior de color
    head = FancyBboxPatch((0.02, 0.78), 0.96, 0.17,
                          boxstyle="round,pad=0.0,rounding_size=0.04",
                          fc=c_fen, ec=c_fen, lw=0, alpha=0.92,
                          transform=ax3.transAxes, zorder=2)
    ax3.add_patch(head)
    ax3.text(0.5, 0.866, "FENOTIPO VASCULAR ÚNICO",
             ha="center", va="center", fontsize=8.2, fontweight="bold",
             color="white", transform=ax3.transAxes, zorder=3)

    ax3.text(0.5, 0.58, fenotipo_graf, ha="center", va="center",
             fontsize=28, fontweight="bold", color=c_fen, transform=ax3.transAxes,
             zorder=3)
    ax3.text(0.5, 0.41, sub, ha="center", va="center",
             fontsize=8.6, color="#334155", transform=ax3.transAxes, zorder=3)
    ax3.text(0.5, 0.27, rango_txt, ha="center", va="center",
             fontsize=9, fontweight="bold", color=c_fen, transform=ax3.transAxes, zorder=3)
    ax3.text(0.5, 0.16, f"VOP CF medida: {vop} m/s", ha="center", va="center",
             fontsize=8.6, color="#0F172A", transform=ax3.transAxes, zorder=3)
    if vop > 10:
        ax3.text(0.5, 0.07, "⚠ Riesgo CV elevado por VOP > 10 m/s",
                 ha="center", va="center", fontsize=7.5, fontweight="bold",
                 color=COLOR_EVA, transform=ax3.transAxes, zorder=3)

    # ============== Panel 3: percentiles (barras horizontales) ==============
    ax2 = fig.add_subplot(gs[1, 0])
    labels = ["p10", "p25", "p50", "p75", "p90", "Paciente"]
    valores = [p10, p25, p50, p75, p90, vop]
    colores = [COLOR_SUPERNOVA, "#5DADE2", COLOR_NORMAL, "#E59866", COLOR_EVA, color_paciente]
    bars = ax2.barh(labels, valores, color=colores, edgecolor="white", linewidth=1.4, height=0.7)
    for bar, val, lbl in zip(bars, valores, labels):
        peso = "bold" if lbl == "Paciente" else "normal"
        ax2.text(val + 0.14, bar.get_y() + bar.get_height() / 2,
                 f"{val}", va="center", fontsize=8, fontweight=peso, color="#0F172A")
    ax2.set_title("Comparativa por percentiles", color="#0F172A")
    ax2.set_xlabel("VOP (m/s)")
    ax2.invert_yaxis()
    ax2.set_xlim(0, max(valores) * 1.30)
    ax2.grid(True, axis="x", ls=":", alpha=0.35, color="#94A3B8")

    # ============== Panel 4: edad cronológica vs vascular ==============
    ax4 = fig.add_subplot(gs[1, 1])
    delta = edad_vasc - edad
    if delta > 0:
        c_edad = COLOR_EVA; delta_txt = f"+{delta} años"
    elif delta < 0:
        c_edad = COLOR_SUPERNOVA; delta_txt = f"{delta} años"
    else:
        c_edad = COLOR_NORMAL; delta_txt = "0 años"

    barras = ax4.bar(["Cronológica", "Vascular"], [edad, edad_vasc],
                     color=["#475569", c_edad], width=0.55, edgecolor="white", linewidth=1.6)
    ymax = max(edad, edad_vasc, 30) + 22
    ax4.set_ylim(0, ymax)
    ax4.set_ylabel("Años")
    ax4.set_title("Edad cronológica vs edad vascular", color="#0F172A")
    ax4.grid(True, axis="y", ls=":", alpha=0.35, color="#94A3B8")
    for bar, val in zip(barras, [edad, edad_vasc]):
        ax4.text(bar.get_x() + bar.get_width() / 2, val + 1.5, f"{val} a",
                 ha="center", fontsize=9.5, fontweight="bold", color="#0F172A")
    # Caja delta (coords en axes 0..1)
    ax4.text(0.5, 0.92, f"$\\Delta$: {delta_txt}", ha="center", va="center",
             fontsize=10.5, fontweight="bold", color="white",
             transform=ax4.transAxes,
             bbox=dict(boxstyle="round,pad=0.35", fc=c_edad, ec="none"))

    # ============== Panel 5: gauge de riesgo cardiovascular ==============
    ax5 = fig.add_subplot(gs[1, 2])
    ax5.axis("off")
    ax5.set_xlim(-1.2, 1.2); ax5.set_ylim(-0.2, 1.25)
    # Semicírculo dividido en 4 segmentos
    seg_colors = [COLOR_NORMAL, "#F4D03F", "#E67E22", COLOR_EVA]
    seg_labels = ["Bajo", "Moderado", "Alto", "Muy Alto"]
    angles = [180, 135, 90, 45, 0]
    for i in range(4):
        w = Wedge((0, 0), 1.0, angles[i+1], angles[i], width=0.30,
                  fc=seg_colors[i], ec="white", lw=2)
        ax5.add_patch(w)
        a = np.deg2rad((angles[i] + angles[i+1]) / 2)
        ax5.text(0.83 * np.cos(a), 0.83 * np.sin(a), seg_labels[i],
                 ha="center", va="center", fontsize=7.5, fontweight="bold", color="white")

    riesgo_actual = riesgo or "Bajo"
    idx = {"Bajo": 0, "Moderado": 1, "Alto": 2, "Muy Alto": 3}.get(riesgo_actual, 0)
    aguja_ang = np.deg2rad((angles[idx] + angles[idx+1]) / 2)
    ax5.annotate("", xy=(0.78 * np.cos(aguja_ang), 0.78 * np.sin(aguja_ang)),
                 xytext=(0, 0),
                 arrowprops=dict(arrowstyle="-|>", color="#0F172A", lw=2.2))
    ax5.scatter(0, 0, s=80, color="#0F172A", zorder=5)
    ax5.set_title("Riesgo cardiovascular global", color="#0F172A", pad=6)
    ax5.text(0, -0.12, riesgo_actual, ha="center", va="top",
             fontsize=12, fontweight="bold", color=seg_colors[idx])

    # ============== Panel 6: franja informativa hemodinámica ==============
    ax6 = fig.add_subplot(gs[2, :])
    ax6.axis("off")
    metrics = []
    if pas and pad:
        pp = pas - pad
        pam = round(pad + (pas - pad) / 3, 1)
        metrics = [
            ("PAS", f"{pas} mmHg", COLOR_EVA if pas >= 140 else COLOR_NORMAL),
            ("PAD", f"{pad} mmHg", COLOR_EVA if pad >= 90 else COLOR_NORMAL),
            ("PP",  f"{pp} mmHg",   COLOR_EVA if pp >= 60 else (COLOR_NORMAL if pp < 50 else "#E67E22")),
            ("PAM", f"{pam} mmHg",  COLOR_NORMAL),
            ("VOP CF", f"{vop} m/s", color_paciente),
            ("Edad vascular", f"{edad_vasc} a", color_paciente),
        ]
    else:
        metrics = [
            ("VOP CF", f"{vop} m/s", color_paciente),
            ("Edad vascular", f"{edad_vasc} a", color_paciente),
            ("p50", f"{p50} m/s", COLOR_NORMAL),
        ]
    n = len(metrics)
    w_box = 1.0 / n
    for i, (lbl, val, col) in enumerate(metrics):
        x0 = i * w_box + 0.005
        box = FancyBboxPatch((x0, 0.10), w_box - 0.01, 0.80,
                             boxstyle="round,pad=0.01,rounding_size=0.06",
                             fc="white", ec=col, lw=1.6, transform=ax6.transAxes)
        ax6.add_patch(box)
        ax6.text(x0 + w_box / 2 - 0.005, 0.66, lbl, ha="center", va="center",
                 fontsize=8, color="#475569", fontweight="bold", transform=ax6.transAxes)
        ax6.text(x0 + w_box / 2 - 0.005, 0.36, val, ha="center", va="center",
                 fontsize=12.5, color=col, fontweight="bold", transform=ax6.transAxes)

    # Suptítulo elegante con barra acento
    fig.suptitle("Lámina Didáctica — Evaluación de Rigidez Arterial",
                 fontsize=14, fontweight="bold", color="#0A2540", y=0.985)
    fig.text(0.5, 0.962, "VOP carotídeo-femoral · Percentiles ARTERY 2024 · Estratificación de riesgo CV",
             ha="center", fontsize=8.5, color="#64748B", style="italic")

    # Forzar bbox exacto a la figura para evitar que algún artista invisible
    # (anotaciones con arcos, halos con stroke effects, legendas con frame)
    # haga que bbox_inches="tight" infle el bounding box detectado.
    import matplotlib.transforms as mtransforms
    fig.subplots_adjust(left=0.06, right=0.97, top=0.93, bottom=0.04)
    buf = BytesIO()
    # bbox explícito: figura completa, sin tight.
    bbox = mtransforms.Bbox.from_bounds(0, 0, fig.get_figwidth(), fig.get_figheight())
    fig.savefig(buf, format="png", dpi=180, facecolor="white", bbox_inches=bbox)
    plt.close(fig)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# GRÁFICO PLOTLY INTERACTIVO (UI en pantalla)
# ---------------------------------------------------------------------------
def construir_grafico_plotly(edad, sexo, vop, p10, p25, p50, p75, p90,
                             color_paciente, edad_vasc, pas=None, pad=None,
                             riesgo="Bajo"):
    """Lámina interactiva en Plotly con 4 paneles para mostrar en Streamlit."""
    if not _PLOTLY_OK:
        return None

    fig = make_subplots(
        rows=2, cols=2,
        specs=[[{"type": "xy"}, {"type": "bar"}],
               [{"type": "bar"}, {"type": "indicator"}]],
        subplot_titles=(
            f"<b>Curva de rigidez arterial — Sexo {sexo}</b>",
            "<b>Comparativa por percentiles</b>",
            "<b>Edad cronológica vs edad vascular</b>",
            "<b>Riesgo cardiovascular global</b>",
        ),
        column_widths=[0.58, 0.42],
        row_heights=[0.55, 0.45],
        horizontal_spacing=0.10,
        vertical_spacing=0.18,
    )

    # ---- Panel 1: Curva de rigidez ----
    edades = np.linspace(20, 80, 200)
    base = 5.5 if sexo == "Femenino" else 6.0
    cp50 = base + 0.08 * (edades - 20)
    cp10 = cp50 * 0.80
    cp90 = cp50 * 1.30
    cp_max = cp90 * 1.6

    # Bandas (relleno) - SUPERNOVA, HVA, EVA
    fig.add_trace(go.Scatter(x=edades, y=cp10, fill="tozeroy",
                             line=dict(color=COLOR_SUPERNOVA, width=1.4, dash="dash"),
                             fillcolor="rgba(31,111,178,0.16)",
                             name="SUPERNOVA (< p10)",
                             hovertemplate="Edad %{x:.0f} a<br>p10: %{y:.2f} m/s<extra></extra>"),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=edades, y=cp90, fill="tonexty",
                             line=dict(color="#1F2937", width=2.2),
                             fillcolor="rgba(30,132,73,0.14)",
                             name="HVA (p10 - p90)",
                             hovertemplate="Edad %{x:.0f} a<br>p90: %{y:.2f} m/s<extra></extra>"),
                  row=1, col=1)
    fig.add_trace(go.Scatter(x=edades, y=cp_max, fill="tonexty",
                             line=dict(color=COLOR_EVA, width=1.4, dash="dash"),
                             fillcolor="rgba(192,57,43,0.18)",
                             name="EVA (> p90)",
                             hovertemplate="Edad %{x:.0f} a<extra></extra>"),
                  row=1, col=1)

    # Línea umbral 10 m/s
    fig.add_hline(y=10, line=dict(color="#8B0000", width=1.4, dash="dot"),
                  annotation_text="Umbral 10 m/s — Riesgo CV",
                  annotation_position="top left",
                  annotation_font=dict(color="#8B0000", size=10),
                  row=1, col=1)

    # Marcador del paciente con halo
    fig.add_trace(go.Scatter(
        x=[edad], y=[vop],
        mode="markers+text",
        marker=dict(size=24, color=color_paciente, line=dict(color="white", width=3),
                    symbol="circle"),
        text=[f"<b>{vop} m/s</b>"],
        textposition="top center",
        textfont=dict(color=color_paciente, size=11),
        name=f"Paciente ({edad} a)",
        hovertemplate=f"<b>Paciente</b><br>Edad: {edad} a<br>VOP: {vop} m/s<extra></extra>",
    ), row=1, col=1)

    fig.update_xaxes(title_text="Edad cronológica (años)", range=[20, 82],
                     row=1, col=1, gridcolor="rgba(148,163,184,0.25)")
    fig.update_yaxes(title_text="VOP CF (m/s)", range=[2.5, max(16, vop + 3.5)],
                     row=1, col=1, gridcolor="rgba(148,163,184,0.25)")

    # ---- Panel 2: Barras percentiles ----
    labels = ["p10", "p25", "p50", "p75", "p90", "Paciente"]
    valores = [p10, p25, p50, p75, p90, vop]
    colores = [COLOR_SUPERNOVA, "#5DADE2", COLOR_NORMAL, "#E59866", COLOR_EVA, color_paciente]
    fig.add_trace(go.Bar(
        x=valores, y=labels, orientation="h",
        marker=dict(color=colores, line=dict(color="white", width=1.5)),
        text=[f"<b>{v}</b>" for v in valores],
        textposition="outside",
        showlegend=False,
        hovertemplate="<b>%{y}</b>: %{x} m/s<extra></extra>",
    ), row=1, col=2)
    fig.update_xaxes(title_text="VOP (m/s)", row=1, col=2,
                     gridcolor="rgba(148,163,184,0.25)",
                     range=[0, max(valores) * 1.30])
    fig.update_yaxes(autorange="reversed", row=1, col=2)

    # ---- Panel 3: Edad cronológica vs vascular ----
    delta = edad_vasc - edad
    if delta > 0:
        c_delta = COLOR_EVA; delta_txt = f"+{delta} años"
    elif delta < 0:
        c_delta = COLOR_SUPERNOVA; delta_txt = f"{delta} años"
    else:
        c_delta = COLOR_NORMAL; delta_txt = "0 años"
    fig.add_trace(go.Bar(
        x=["Cronológica", "Vascular"],
        y=[edad, edad_vasc],
        marker=dict(color=["#475569", c_delta], line=dict(color="white", width=2)),
        text=[f"<b>{edad} a</b>", f"<b>{edad_vasc} a</b>"],
        textposition="outside",
        showlegend=False,
        hovertemplate="%{x}: %{y} años<extra></extra>",
    ), row=2, col=1)
    fig.update_yaxes(title_text="Años", row=2, col=1,
                     gridcolor="rgba(148,163,184,0.25)",
                     range=[0, max(edad, edad_vasc) + 25])
    fig.add_annotation(
        x=0.5, y=max(edad, edad_vasc) + 15,
        xref=f"x3", yref="y3",
        text=f"<b>Δ: {delta_txt}</b>",
        showarrow=False,
        font=dict(color="white", size=12),
        bgcolor=c_delta, borderpad=6,
        bordercolor=c_delta, borderwidth=1,
    )

    # ---- Panel 4: Gauge de riesgo ----
    riesgo_map = {"Bajo": 1, "Moderado": 2, "Alto": 3, "Muy Alto": 4}
    riesgo_val = riesgo_map.get(riesgo, 1)
    riesgo_color = {"Bajo": COLOR_NORMAL, "Moderado": "#F4D03F",
                    "Alto": "#E67E22", "Muy Alto": COLOR_EVA}.get(riesgo, COLOR_NORMAL)
    fig.add_trace(go.Indicator(
        mode="gauge+number",
        value=riesgo_val,
        number=dict(suffix=f" — {riesgo}", font=dict(size=18, color=riesgo_color)),
        gauge=dict(
            axis=dict(range=[0, 4], tickwidth=1, tickcolor="#475569",
                      tickvals=[0.5, 1.5, 2.5, 3.5],
                      ticktext=["Bajo", "Moderado", "Alto", "Muy Alto"]),
            bar=dict(color=riesgo_color, thickness=0.32),
            steps=[
                dict(range=[0, 1], color="rgba(30,132,73,0.30)"),
                dict(range=[1, 2], color="rgba(244,208,63,0.30)"),
                dict(range=[2, 3], color="rgba(230,126,34,0.30)"),
                dict(range=[3, 4], color="rgba(192,57,43,0.30)"),
            ],
            threshold=dict(line=dict(color="#0F172A", width=4),
                           thickness=0.85, value=riesgo_val),
        ),
        domain=dict(x=[0, 1], y=[0, 1]),
    ), row=2, col=2)

    fig.update_layout(
        height=720,
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.12,
                    xanchor="center", x=0.30, font=dict(size=10)),
        font=dict(family="Inter, Segoe UI, sans-serif", color="#0F172A", size=11),
        plot_bgcolor="rgba(248,250,252,0.55)",
        paper_bgcolor="rgba(0,0,0,0)",
        margin=dict(l=50, r=30, t=70, b=30),
        title=dict(
            text="<b>Lámina Interactiva — Análisis Vascular Integral</b>",
            x=0.5, xanchor="center",
            font=dict(size=16, color="#0A2540"),
        ),
    )
    return fig


# ---------------------------------------------------------------------------
# GENERADOR DE PDF PROFESIONAL (sin cambios funcionales)
# ---------------------------------------------------------------------------
class PDFReport(FPDF):
    """PDF integrado en 3 hojas A4.

    Hoja 1: datos del paciente, resultados, diagnóstico y captura carótido-femoral.
    Hoja 2: lámina didáctica e interpretación clínica.
    Hoja 3: recomendaciones, cierre médico, firma y sello digital con espacio suficiente.
    """
    def __init__(self, profesional="", firma_path=None, sello_path=None):
        super().__init__(orientation="P", unit="mm", format="A4")
        self.profesional = profesional or "Profesional Responsable"
        self.firma_path = firma_path
        self.sello_path = sello_path
        self.set_auto_page_break(auto=True, margin=14)
        self.set_margins(left=12, top=10, right=12)
        self.alias_nb_pages()

    def header(self):
        self.set_fill_color(*COLOR_BG_HEADER)
        self.rect(0, 0, 210, 18, "F")
        self.set_text_color(255, 255, 255)
        self.set_font("Arial", "B", 11.5)
        self.set_xy(12, 4)
        self.cell(0, 5, safe_latin1("LABORATORIO VASCULAR NO INVASIVO"), 0, 1, "L")
        self.set_font("Arial", "I", 7.8)
        self.set_x(12)
        self.cell(0, 4, safe_latin1("Evaluación de Rigidez Arterial y Fenotipado Vascular - EVA/SUPERNOVA"), 0, 1, "L")
        self.set_xy(150, 4)
        self.set_font("Arial", "", 7.2)
        self.cell(48, 4, safe_latin1(f"Folio: VAS-{datetime.datetime.now().strftime('%Y%m%d%H%M')}"), 0, 1, "R")
        self.set_x(150)
        self.cell(48, 4, safe_latin1(f"Fecha: {datetime.date.today().strftime('%d/%m/%Y')}"), 0, 1, "R")
        self.set_text_color(0, 0, 0)
        self.set_y(22)

    def footer(self):
        self.set_y(-12)
        self.set_draw_color(*COLOR_BG_HEADER)
        self.set_line_width(0.25)
        self.line(12, self.get_y(), 198, self.get_y())
        self.set_font("Arial", "I", 6.8)
        self.set_text_color(90, 90, 90)
        self.cell(0, 3.6, safe_latin1("Documento de uso clínico. Interpretación por personal médico calificado."), 0, 1, "C")
        self.cell(0, 3.6, safe_latin1(f"Página {self.page_no()} / {{nb}}"), 0, 0, "C")
        self.set_text_color(0, 0, 0)

    def section_title(self, t):
        self.set_fill_color(*COLOR_BG_HEADER)
        self.set_text_color(255, 255, 255)
        self.set_font("Arial", "B", 8.8)
        self.cell(0, 5.4, safe_latin1("  " + t.upper()), 0, 1, "L", fill=True)
        self.set_text_color(0, 0, 0)
        self.ln(0.8)

    def patient_info(self, datos):
        self.section_title("Datos del Paciente")
        self.set_fill_color(*COLOR_BG_SECTION)
        filas = [
            ("Nombre", datos.get("nombre", "-"), "Documento", datos.get("documento", "-")),
            ("Edad", f"{datos.get('edad', '-')} años", "Sexo", datos.get("sexo", "-")),
            ("Médico solicitante", datos.get("medico_solicitante", "-"),
             "Fecha del estudio", datetime.date.today().strftime("%d/%m/%Y")),
        ]
        for f in filas:
            self.set_font("Arial", "B", 8.2)
            self.cell(34, 5.8, safe_latin1(f[0]), 1, 0, "L", fill=True)
            self.set_font("Arial", "", 8.2)
            self.cell(60, 5.8, safe_latin1(str(f[1])[:42]), 1, 0, "L")
            self.set_font("Arial", "B", 8.2)
            self.cell(39, 5.8, safe_latin1(f[2]), 1, 0, "L", fill=True)
            self.set_font("Arial", "", 8.2)
            self.cell(53, 5.8, safe_latin1(str(f[3])[:38]), 1, 1, "L")
        self.ln(2)

    def resultados_table(self, r):
        self.section_title("Resultados Hemodinámicos")
        self.set_font("Arial", "B", 8.1)
        self.set_fill_color(*COLOR_BG_HEADER)
        self.set_text_color(255, 255, 255)
        for h, w in [("Parámetro", 68), ("Resultado", 44), ("Referencia", 74)]:
            self.cell(w, 5.4, safe_latin1(h), 1, 0, "C", fill=True)
        self.ln()
        self.set_text_color(0, 0, 0)
        filas = [
            ("VOP carótido-femoral", f"{r['vop']} m/s", f"p50: {r['p50']} m/s"),
            ("Presión Arterial", f"{r['pas']}/{r['pad']} mmHg", "< 130/80 mmHg"),
            ("Presión de Pulso", f"{r['pp']} mmHg", "< 50 mmHg"),
            ("Presión Arterial Media", f"{r['pam']} mmHg", "70 - 105 mmHg"),
            ("Percentil 10", f"{r['p10']} m/s", "SUPERNOVA si VOP < p10"),
            ("Percentil 90", f"{r['p90']} m/s", "EVA si VOP > p90"),
            ("Lesión Órgano Blanco", r["lob"], "VOP > 10 m/s"),
            ("Riesgo CV global", r["riesgo"], "Estratificación clínica"),
        ]
        fill = False
        for f in filas:
            self.set_fill_color(245, 247, 250)
            self.set_font("Arial", "", 7.9)
            self.cell(68, 5.2, safe_latin1(f[0]), 1, 0, "L", fill=fill)
            self.set_font("Arial", "B", 7.9)
            self.cell(44, 5.2, safe_latin1(f[1]), 1, 0, "C", fill=fill)
            self.set_font("Arial", "", 7.9)
            self.cell(74, 5.2, safe_latin1(f[2]), 1, 1, "L", fill=fill)
            fill = not fill
        self.ln(2)

    def diagnostico_box(self, fenotipo, riesgo, vop=None):
        self.section_title("Conclusión Diagnóstica")
        if "EVA" in fenotipo:
            fill = COLOR_BG_ALERT; txt = (155, 30, 20)
        elif "SUPERNOVA" in fenotipo:
            fill = (217, 232, 252); txt = (28, 80, 140)
        else:
            fill = COLOR_BG_OK; txt = (28, 110, 60)
        tiene_lob = vop is not None and vop > 10
        alto = 18 if tiene_lob else 14
        y0 = self.get_y()
        self.set_fill_color(*fill)
        self.set_draw_color(*txt)
        self.set_text_color(*txt)
        self.set_line_width(0.3)
        self.rect(12, y0, 186, alto, "DF")
        self.set_xy(15, y0 + 2.5)
        self.set_font("Arial", "B", 9.8)
        self.cell(0, 5, safe_latin1(f"Fenotipo vascular único: {fenotipo}"), 0, 1, "L")
        self.set_x(15)
        self.set_font("Arial", "B", 8.6)
        if tiene_lob:
            self.cell(0, 4.4, safe_latin1(f"Riesgo cardiovascular: {riesgo} - elevado por VOP mayor a 10 m/seg"), 0, 1, "L")
        else:
            self.cell(0, 4.4, safe_latin1(f"Riesgo cardiovascular global: {riesgo}"), 0, 1, "L")
        if tiene_lob:
            self.set_x(15)
            self.set_text_color(155, 30, 20)
            self.set_font("Arial", "B", 8.1)
            self.cell(0, 4.2, safe_latin1(f"LOB por rigidez vascular: VOP {vop} m/s > 10 m/s."), 0, 1, "L")
        self.set_y(y0 + alto + 3)
        self.set_text_color(0, 0, 0)
        self.set_draw_color(0, 0, 0)

    def insert_imported_cf_curve(self, image_bytes):
        if not image_bytes:
            return
        self.section_title("Curva Carótido-Femoral Importada del Estudio Original")
        self.set_font("Arial", "", 7.4)
        self.multi_cell(0, 3.8, safe_latin1(
            "Captura del estudio original. TCF: tiempo de tránsito carótido-femoral. Dist. Car. Fem.: distancia carótido-femoral."
        ))
        self.ln(1)
        bio = BytesIO(image_bytes) if isinstance(image_bytes, (bytes, bytearray)) else image_bytes
        bio.seek(0)
        try:
            if self.get_y() > 190:
                self.add_page()
            self.image(bio, x=22, w=150)
        except Exception:
            import tempfile, os
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tf:
                bio.seek(0)
                tf.write(bio.read())
                tmp_path = tf.name
            try:
                self.image(tmp_path, x=22, w=150)
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        self.ln(2)

    def insert_chart(self, image_bytes):
        self.section_title("Lámina Didáctica - Análisis Gráfico")
        image_bytes.seek(0)
        try:
            self.image(image_bytes, x=14, w=180)
        except Exception:
            import tempfile, os
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tf:
                image_bytes.seek(0)
                tf.write(image_bytes.read())
                tmp_path = tf.name
            try:
                self.image(tmp_path, x=14, w=180)
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        self.ln(2)

    def interpretacion(self, fenotipo, vop, edad, edad_vasc, riesgo):
        self.section_title("Interpretación Clínica")
        self.set_font("Arial", "", 8.6)
        lob_txt = ""
        if vop > 10:
            lob_txt = "\n\nVOP > 10 m/s: lesión de órgano blanco por rigidez arterial."

        if "EVA" in fenotipo:
            def_fen = "Envejecimiento Vascular Acelerado: VOP por encima del percentil 90 para sexo y edad."
        elif "SUPERNOVA" in fenotipo:
            def_fen = "SUPERNOVA / Envejecimiento Vascular Supranormal: VOP carótido-femoral por debajo del percentil 10 para sexo y edad."
        else:
            def_fen = "HVA / Envejecimiento Vascular Saludable: VOP entre el percentil 10 y el percentil 90 para sexo y edad."

        txt = (
            f"La Velocidad de Onda de Pulso (VOP) carótido-femoral medida fue de {vop} m/s. "
            f"El diagnóstico de rigidez vascular se expresa como un fenotipo único, definido por la posición de la VOP frente a percentiles por edad y sexo.\n\n"
            f"Fenotipo vascular único: {fenotipo}.\n"
            f"{def_fen}\n\n"
            f"Riesgo cardiovascular global estimado: {riesgo}."
            f"{lob_txt}"
        )
        self.multi_cell(0, 4.6, safe_latin1(txt))
        self.ln(2)

    def recomendaciones_clinicas(self, recs):
        self.section_title("Recomendaciones Clínicas")
        self.set_font("Arial", "", 8.4)
        for i, r in enumerate((recs or [])[:7], 1):
            self.set_x(12)
            self.multi_cell(0, 4.3, safe_latin1(f"{i}. {r}"))
        self.ln(2)

    def cierre_medico(self, fenotipo, riesgo, vop=None):
        self.section_title("Cierre Médico Integrado")
        self.set_font("Arial", "", 8.6)
        if vop is not None and vop > 10:
            riesgo_txt = (
                f"Riesgo cardiovascular global consignado: {riesgo}. "
                f"El riesgo cardiovascular se considera elevado por VOP mayor a 10 m/seg."
            )
        else:
            riesgo_txt = f"Riesgo cardiovascular global consignado: {riesgo}."
        txt = (
            f"Informe integrado de rigidez vascular con fenotipo final: {fenotipo}. "
            f"La conducta clínica debe integrarse con presión arterial, antecedentes, daño de órgano blanco y criterio médico tratante. "
            f"{riesgo_txt}"
        )
        self.multi_cell(0, 4.6, safe_latin1(txt))
        self.ln(4)

    def firma(self):
        """Firma y sello con espacio reservado en página 3."""
        if self.get_y() > 205:
            self.set_y(205)
        else:
            self.ln(8)
        y0 = self.get_y()

        self.set_draw_color(210, 210, 210)
        self.set_fill_color(250, 250, 250)
        self.rect(108, y0, 88, 48, "D")

        try:
            if self.firma_path and os.path.exists(self.firma_path):
                self.image(self.firma_path, x=114, y=y0 + 4, w=42)
            if self.sello_path and os.path.exists(self.sello_path):
                self.image(self.sello_path, x=160, y=y0 + 3, w=30)
            if self.firma_path or self.sello_path:
                self.set_y(y0 + 30)
            else:
                self.set_y(y0 + 27)
        except Exception:
            self.set_y(y0 + 27)

        self.set_draw_color(60, 60, 60)
        self.line(112, self.get_y(), 192, self.get_y())
        self.set_font("Arial", "B", 8.8)
        self.set_xy(112, self.get_y() + 1.5)
        self.cell(80, 4.4, safe_latin1(self.profesional), 0, 1, "C")
        self.set_font("Arial", "I", 7.8)
        self.set_x(112)
        self.cell(80, 4.2, safe_latin1("Médico responsable del estudio"), 0, 1, "C")


def construir_pdf(datos, resultados, fenotipo, riesgo, recs, chart_buf, profesional,
                  curva_cf_png=None, firma_path=None, sello_path=None):
    pdf = PDFReport(profesional=profesional, firma_path=firma_path, sello_path=sello_path)

    # Hoja 1: resumen clínico + evidencia original CF.
    pdf.add_page()
    pdf.patient_info(datos)
    pdf.resultados_table(resultados)
    pdf.diagnostico_box(fenotipo, riesgo, vop=resultados.get("vop"))
    pdf.insert_imported_cf_curve(curva_cf_png)

    # Hoja 2: gráfico didáctico + interpretación.
    pdf.add_page()
    pdf.insert_chart(chart_buf)
    pdf.interpretacion(fenotipo, resultados["vop"], resultados["edad"],
                       resultados["edad_vasc"], riesgo)

    # Hoja 3: recomendaciones + cierre + firma y sello con espacio suficiente.
    pdf.add_page()
    pdf.recomendaciones_clinicas(recs)
    pdf.cierre_medico(fenotipo, riesgo, vop=resultados.get("vop"))
    pdf.firma()

    # Seguridad: limitar el informe integrado a 3 páginas.
    try:
        while getattr(pdf, "page", 0) > 3:
            pdf.pages.pop(pdf.page, None)
            pdf.page -= 1
    except Exception:
        pass

    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1")
    return bytes(out)


# ---------------------------------------------------------------------------
# COMPONENTES DE UI MODERNOS
# ---------------------------------------------------------------------------
def _hero(titulo: str, subtitulo: str):
    st.markdown(
        f"""
        <div class="vh-hero">
            <h1>{titulo}</h1>
            <p>{subtitulo}</p>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _card_open(titulo: str):
    st.markdown(
        f'<div class="vh-card"><div class="vh-card-title">{titulo}</div>',
        unsafe_allow_html=True,
    )


def _card_close():
    st.markdown("</div>", unsafe_allow_html=True)


def _badge_fenotipo(fenotipo: str) -> str:
    if "EVA" in fenotipo:
        cls = "vh-badge-eva"
    elif "SUPERNOVA" in fenotipo:
        cls = "vh-badge-supernova"
    else:
        cls = "vh-badge-normal"
    return f'<span class="vh-badge {cls}">{fenotipo}</span>'


# ---------------------------------------------------------------------------
# INTERFAZ STREAMLIT MODERNIZADA
# ---------------------------------------------------------------------------
def _pantalla_login():
    """Pantalla de acceso modernizada con card centrada."""
    st.markdown(
        """
        <div class="vh-login-wrap">
            <div class="vh-login-card">
                <div class="vh-login-header">
                    <div class="vh-login-logo">🫀</div>
                    <h2 class="vh-login-title">Vascular Health Analyzer</h2>
                    <p class="vh-login-sub">Evaluación de rigidez arterial · EVA / SUPERNOVA</p>
                </div>
        """,
        unsafe_allow_html=True,
    )

    if _USERS is None:
        st.error("No se pudo inicializar la gestión de usuarios.")
        st.code(f"Detalle del error: {_USER_ERR}")
        st.markdown(
            "**Diagnóstico rápido:**\n\n"
            "1. Verifique que `users.py` esté en la misma carpeta que `app.py`.\n"
            "2. Verifique su versión de Python: requiere **Python 3.8 o superior**.\n"
            "3. Si la carpeta es de solo lectura (ej. Google Drive sincronizado), "
            "mueva `app.py` y `users.py` a una carpeta local con permisos de escritura."
        )
        st.markdown("</div></div>", unsafe_allow_html=True)
        return

    tab_login, tab_reg, tab_rec = st.tabs(
        ["🔐 Iniciar sesión", "✨ Registrarse", "🔄 Recuperar contraseña"]
    )

    with tab_login:
        user = st.text_input("Usuario", key="login_user", placeholder="su.usuario")
        pw = st.text_input("Contraseña", type="password", key="login_pw",
                           placeholder="••••••••")
        if st.button("Ingresar", key="btn_login", use_container_width=True):
            ok, role_or_msg = _USERS.verificar(user, pw)
            if ok:
                st.session_state.auth = True
                st.session_state.user_role = role_or_msg
                st.session_state.username = user.strip().lower()
                st.rerun()
            else:
                st.error(role_or_msg)

    with tab_reg:
        st.caption("Cree una cuenta médica nueva. La pregunta de seguridad le servirá "
                   "para recuperar la contraseña.")
        col_a, col_b = st.columns(2)
        with col_a:
            r_user = st.text_input("Nuevo usuario (alfanum., mín. 3)", key="reg_user")
            r_nombre = st.text_input("Nombre completo", key="reg_nombre")
            r_pw = st.text_input("Contraseña (mín. 8)", type="password", key="reg_pw")
            r_q = st.text_input("Pregunta de seguridad",
                                placeholder="Ej.: Nombre de su primera mascota",
                                key="reg_q")
        with col_b:
            r_matricula = st.text_input("Matrícula profesional (opcional)",
                                        key="reg_matricula")
            r_pw2 = st.text_input("Repita la contraseña", type="password", key="reg_pw2")
            r_a = st.text_input("Respuesta a la pregunta", key="reg_a")

        st.markdown("**Firma y sello digital (opcional)**")
        st.caption("Estos archivos quedan asociados exclusivamente a este usuario y solo "
                   "se insertan en sus propios informes PDF.")
        col_f, col_s = st.columns(2)
        with col_f:
            r_firma = st.file_uploader("Firma digital", type=["png", "jpg", "jpeg"],
                                       key="reg_firma")
        with col_s:
            r_sello = st.file_uploader("Sello digital", type=["png", "jpg", "jpeg"],
                                       key="reg_sello")

        if st.button("Crear cuenta", key="btn_reg", use_container_width=True):
            if r_pw != r_pw2:
                st.error("Las contraseñas no coinciden.")
            else:
                ok, msg = _USERS.registrar(
                    r_user, r_pw, r_q, r_a,
                    role="medico",
                    nombre_completo=r_nombre,
                    matricula=r_matricula,
                )
                if ok:
                    try:
                        _guardar_asset_usuario(r_user, r_firma, "firma")
                        _guardar_asset_usuario(r_user, r_sello, "sello")
                    except Exception as exc:
                        st.warning(f"La cuenta fue creada, pero no se pudo guardar firma/sello: {exc}")
                    st.success(msg)
                else:
                    st.error(msg)

    with tab_rec:
        st.caption("Ingrese su usuario para ver su pregunta de seguridad. "
                   "Luego responda y elija una contraseña nueva.")
        rc_user = st.text_input("Usuario", key="rec_user")
        if rc_user and _USERS.existe(rc_user):
            pregunta = _USERS.obtener_pregunta(rc_user)
            st.info(f"Pregunta de seguridad: **{pregunta}**")
            rc_a = st.text_input("Su respuesta", key="rec_a")
            rc_pw = st.text_input("Nueva contraseña (mín. 8)",
                                  type="password", key="rec_pw")
            rc_pw2 = st.text_input("Repita la nueva contraseña",
                                   type="password", key="rec_pw2")
            if st.button("Restablecer contraseña", key="btn_rec",
                         use_container_width=True):
                if rc_pw != rc_pw2:
                    st.error("Las contraseñas no coinciden.")
                else:
                    ok, msg = _USERS.recuperar(rc_user, rc_a, rc_pw)
                    if ok:
                        st.success(msg)
                    else:
                        st.error(msg)
        elif rc_user:
            st.warning("Usuario no encontrado.")

    st.markdown("</div></div>", unsafe_allow_html=True)


def _sidebar(usuario_actual: str):
    """Sidebar modernizada con perfil, firma/sello y navegación."""
    rol = st.session_state.get("user_role", "medico").capitalize()
    st.sidebar.markdown(
        f"""
        <div style="padding:18px 8px 12px 8px; text-align:center;">
            <div style="width:60px;height:60px;margin:0 auto 8px auto;
                        border-radius:16px;background:linear-gradient(135deg,#00B4D8,#0077B6);
                        display:flex;align-items:center;justify-content:center;
                        font-size:28px;color:white;
                        box-shadow:0 8px 22px rgba(0,180,216,0.4);">
                {usuario_actual[:1].upper() if usuario_actual else '?'}
            </div>
            <div style="color:#FFFFFF;font-weight:700;font-size:15px;">
                {usuario_actual or '—'}
            </div>
            <div style="color:#94A3B8;font-size:11px;text-transform:uppercase;
                        letter-spacing:0.6px;margin-top:2px;">
                {rol}
            </div>
        </div>
        <hr style="border:none;border-top:1px solid rgba(255,255,255,0.08);margin:6px 0 14px 0;">
        """,
        unsafe_allow_html=True,
    )

    profesional = st.sidebar.text_input(
        "Profesional responsable",
        value=st.session_state.get("profesional_default",
                                   "Dr. / Dra. ____________________"),
        key="sidebar_profesional",
    )

    perfil_actual = _perfil_usuario_actual(usuario_actual)
    with st.sidebar.expander("✍ Firma y sello digital"):
        col_a, col_b = st.columns(2)
        with col_a:
            if perfil_actual.get("firma_path"):
                st.image(perfil_actual["firma_path"], caption="Firma", use_container_width=True)
            else:
                st.caption("Sin firma")
        with col_b:
            if perfil_actual.get("sello_path"):
                st.image(perfil_actual["sello_path"], caption="Sello", use_container_width=True)
            else:
                st.caption("Sin sello")

        up_firma = st.file_uploader("Actualizar firma", type=["png", "jpg", "jpeg"],
                                    key="up_firma_usuario")
        up_sello = st.file_uploader("Actualizar sello", type=["png", "jpg", "jpeg"],
                                    key="up_sello_usuario")
        if st.button("Guardar firma/sello", key="btn_guardar_assets_usuario",
                     use_container_width=True):
            try:
                if up_firma is not None:
                    _guardar_asset_usuario(usuario_actual, up_firma, "firma")
                if up_sello is not None:
                    _guardar_asset_usuario(usuario_actual, up_sello, "sello")
                st.success("Firma/sello guardados.")
                st.rerun()
            except Exception as exc:
                st.error(f"No se pudo guardar firma/sello: {exc}")

    menu = ["🩺 Nuevo Estudio", "📊 Historial y Exportación"]
    if st.session_state.get("user_role") == "admin":
        menu.append("👥 Administrar Usuarios")
    choice = st.sidebar.selectbox("Menú", menu, key="sidebar_menu")

    st.sidebar.markdown(
        "<hr style='border:none;border-top:1px solid rgba(255,255,255,0.08);"
        "margin:14px 0 10px 0;'>",
        unsafe_allow_html=True,
    )
    if st.sidebar.button("🚪 Cerrar sesión", key="btn_logout",
                         use_container_width=True):
        st.session_state.auth = False
        st.rerun()

    st.sidebar.caption(
        f"Plotly: {'✓ disponible' if _PLOTLY_OK else '— no instalado'}"
    )

    return profesional, choice


def _pantalla_admin():
    st.markdown('<div class="vh-card">', unsafe_allow_html=True)
    st.markdown("### 👥 Administración de Usuarios")
    if _USERS is None:
        st.error("Gestión de usuarios no disponible.")
        st.markdown("</div>", unsafe_allow_html=True)
        return
    st.markdown("#### Usuarios registrados")
    st.write(_USERS.lista_usuarios())

    st.markdown("#### Resetear contraseña de un usuario")
    target = st.selectbox("Usuario objetivo", _USERS.lista_usuarios(),
                          key="adm_target")
    col_a, col_b = st.columns(2)
    with col_a:
        adm_pw = st.text_input("Su contraseña de admin (confirmación)",
                               type="password", key="adm_pw")
    with col_b:
        new_pw = st.text_input("Nueva contraseña (mín. 8 caracteres)",
                               type="password", key="adm_newpw")
    if st.button("Aplicar reset", key="btn_admreset"):
        ok, msg = _USERS.admin_reset(
            st.session_state.get("username", "admin"),
            adm_pw, target, new_pw,
        )
        if ok:
            st.success(msg)
        else:
            st.error(msg)
    st.markdown("</div>", unsafe_allow_html=True)


def _pantalla_nuevo_estudio(profesional: str):
    _hero("Registro de Evaluación Vascular",
          "Importe el informe original (PDF/TXT) o complete los datos manualmente. "
          "Se genera diagnóstico, lámina interactiva e informe PDF profesional.")

    # ====== Importador ======
    with st.expander("📤 Importar mediciones desde archivo PDF o TXT", expanded=True):
        st.caption("Suba el informe PDF o el archivo TXT sin formato del equipo de medición. "
                   "Los campos se completarán automáticamente cuando el texto sea reconocible.")
        pdf_in = st.file_uploader("Seleccionar archivo PDF o TXT",
                                  type=["pdf", "txt"], key="importacion_upload")
        if pdf_in is not None:
            es_txt = str(getattr(pdf_in, "name", "")).lower().endswith(".txt")
            if _PDF_BACKEND is None and not es_txt:
                st.error("Falta lector de PDF. Instale: `pip install pdfplumber`. "
                         "Los archivos TXT pueden importarse sin lector PDF.")
            else:
                try:
                    new_fid = getattr(pdf_in, "file_id", pdf_in.name)
                    if st.session_state.get("importacion_fid") != new_fid:
                        with st.spinner("Analizando archivo y extrayendo mediciones..."):
                            datos = GenericReportParser.parsear(pdf_in)
                        st.session_state["importacion_datos"] = datos
                        st.session_state["importacion_fid"] = new_fid
                        st.session_state["archivo_importado_nombre"] = pdf_in.name
                        mapeo = {
                            "nombre": "f_nombre",
                            "documento": "f_documento",
                            "edad": "f_edad",
                            "sexo": "f_sexo",
                            "pas": "f_pas",
                            "pad": "f_pad",
                            "distancia": "f_distancia",
                            "tiempo": "f_tiempo",
                            "vop": "f_vop_medida",
                        }
                        vacios = {
                            "f_nombre": "", "f_documento": "", "f_edad": 50,
                            "f_sexo": "Masculino", "f_pas": 0, "f_pad": 0,
                            "f_distancia": 0.0, "f_tiempo": 0.0, "f_vop_medida": 0.0,
                        }
                        for kform, v0 in vacios.items():
                            st.session_state[kform] = v0
                        for kparser, kform in mapeo.items():
                            v = datos.get(kparser)
                            if v is None or v == "":
                                continue
                            st.session_state[kform] = v
                        st.rerun()

                    datos = st.session_state["importacion_datos"]
                    detectados = {k: v for k, v in datos.items()
                                  if not k.startswith("_") and v is not None}
                    if detectados:
                        st.success(f"✓ Se detectaron {len(detectados)} campos del archivo. "
                                   "Verifique antes de generar el diagnóstico.")
                        with st.expander("Ver campos detectados (JSON)"):
                            st.json(detectados)
                        todos = ["nombre","documento","edad","sexo","pas","pad",
                                 "distancia","tiempo","vop"]
                        faltantes = [c for c in todos if not datos.get(c)]
                        if faltantes:
                            st.warning(
                                "Campos no detectados automáticamente (complételos "
                                "manualmente abajo): " + ", ".join(faltantes)
                            )
                        st.info("Criterio aplicado: la VOP de referencia es la "
                                "carotídeo-femoral/cfPWV. La VOP radial o periférica "
                                "se ignora para el diagnóstico.")
                    else:
                        st.warning("No se reconocieron campos. Verifique manualmente.")
                    if datos.get("_curva_cf_png"):
                        st.image(datos["_curva_cf_png"],
                                 caption="Sector carótido-femoral capturado del PDF original",
                                 use_container_width=False)
                    elif str(getattr(pdf_in, "name", "")).lower().endswith(".pdf"):
                        st.warning(
                            "No se pudo capturar automáticamente el sector gráfico "
                            "carótido-femoral. Verifique dependencias: PyMuPDF y Pillow. "
                            "Si el PDF es una imagen no detectable, la app intentará incorporar "
                            "una captura de página completa como respaldo."
                        )
                    st.markdown("**Texto extraído del archivo (primeros 4000 caracteres)**")
                    st.text_area(
                        "Texto extraído",
                        value=datos.get("_texto_crudo", "")[:4000],
                        height=220,
                        key="txt_extraido_importacion",
                        label_visibility="collapsed",
                    )
                except Exception as exc:
                    st.error(f"Error al leer el archivo: {exc}")

    ram = st.session_state.get("importacion_datos", {}) or {}

    defaults = {
        "f_nombre": "", "f_documento": "", "f_edad": 50,
        "f_sexo": "Masculino", "f_medsol": "",
        "f_pas": 0, "f_pad": 0,
        "f_distancia": 0.0, "f_tiempo": 0.0,
        "f_vop_medida": 0.0,
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)

    # ====== Formulario ======
    _card_open("📋 Datos del paciente y mediciones")
    with st.form("estudio_form", clear_on_submit=False):
        col1, col2 = st.columns(2)
        with col1:
            st.text_input("Nombre Completo del Paciente", key="f_nombre")
            st.text_input("Documento / Identificación", key="f_documento")
            st.number_input("Edad", min_value=1, max_value=120, step=1, key="f_edad")
            st.radio("Sexo", ["Masculino", "Femenino"], key="f_sexo", horizontal=True)
            st.text_input("Médico solicitante", key="f_medsol")
        with col2:
            st.number_input("Presión Sistólica (PAS, mmHg)",
                            min_value=0, max_value=260, step=1, key="f_pas")
            st.number_input("Presión Diastólica (PAD, mmHg)",
                            min_value=0, max_value=160, step=1, key="f_pad")
            st.number_input(
                "Distancia Carótido-Femoral REAL importada (cm) = Dist. Car. Fem.",
                min_value=0.0, max_value=200.0, step=0.5,
                format="%.1f", key="f_distancia")
            st.number_input(
                "Tiempo de Tránsito CF importado (ms) — TCF",
                min_value=0.0, max_value=500.0, step=0.5,
                format="%.1f", key="f_tiempo")
            st.number_input(
                "VOP CARÓTIDO-FEMORAL medida del archivo original (m/s)",
                min_value=0.0, max_value=25.0, step=0.1,
                format="%.2f", key="f_vop_medida",
                help="Dato primario importado desde el archivo original. Debe ser VOP "
                     "carotídeo-femoral/cfPWV; no se usa VOP radial.")
            if ram.get("vop"):
                st.caption(f"VOP cf medida importada del archivo: **{ram['vop']} m/s** (valor primario)")
            elif ram.get("vop_recalculada"):
                st.caption(f"No se detectó VOP medida en el archivo. Respaldo calculado "
                           f"desde distancia/tiempo: **{ram['vop_recalculada']} m/s**")

        submitted = st.form_submit_button("🧪 Calcular y generar diagnóstico",
                                          use_container_width=True)
    _card_close()

    # Recuperar valores
    nombre = st.session_state["f_nombre"]
    documento = st.session_state["f_documento"]
    edad = int(st.session_state["f_edad"])
    sexo = st.session_state["f_sexo"]
    medico_solicitante = st.session_state["f_medsol"]
    pas = int(st.session_state["f_pas"])
    pad = int(st.session_state["f_pad"])
    distancia = float(st.session_state["f_distancia"])
    tiempo = float(st.session_state["f_tiempo"])
    vop_medida = float(st.session_state.get("f_vop_medida", 0.0) or 0.0)

    if submitted:
        with st.expander("🔎 Verificación de datos capturados", expanded=False):
            st.write({
                "edad": edad, "sexo": sexo, "pas": pas, "pad": pad,
                "distancia_cm": distancia, "tiempo_ms": tiempo,
                "vop_medida_cf": vop_medida,
            })
        errores = []
        if distancia <= 0:
            errores.append("No se importó la distancia carótido-femoral real desde el archivo. Revise el archivo original o cargue el valor manualmente.")
        if tiempo <= 0:
            errores.append("No se importó el tiempo de tránsito CF real desde el archivo. Revise el archivo original o cargue el valor manualmente.")
        if pas <= pad:
            errores.append("La PAS debe ser mayor que la PAD. Revise la importación del archivo: PAS/PAD parecen invertidas o mal leídas.")
        if vop_medida and not (3 <= vop_medida <= 25):
            errores.append("La VOP cf medida debe estar entre 3 y 25 m/s.")
        if not (1 <= edad <= 120):
            errores.append("Edad fuera de rango (1-120).")
        if errores:
            for er in errores:
                st.error(er)
            st.stop()

        e = VascularEngine()
        vop_calc = e.calcular_vop(distancia, tiempo)
        if vop_medida > 0:
            vop = round(vop_medida, 2)
            st.info(f"Se usa la VOP cf medida del archivo original: **{vop} m/s**. "
                    f"Control por distancia/tiempo: {vop_calc} m/s")
        elif ram.get("vop_recalculada"):
            vop = float(ram["vop_recalculada"])
            st.warning(f"No se detectó VOP cf medida. Se usa VOP recalculada desde "
                       f"distancia/tiempo: **{vop} m/s**")
        else:
            vop = vop_calc
            st.warning(f"No se detectó VOP cf medida en el archivo. Se usa recálculo "
                       f"desde distancia/tiempo: **{vop} m/s**")
        p10, p25, p50, p75, p90 = e.obtener_percentiles(edad, sexo)
        fenotipo, color, _ = e.clasificar_fenotipo(vop, p10, p90)
        edad_vasc = e.edad_vascular(vop, sexo)
        pp = e.presion_pulso(pas, pad)
        pam = e.presion_arterial_media(pas, pad)
        riesgo = e.riesgo_cv_global(vop, pas, pad, edad)
        lob = "SI (VOP > 10 m/s)" if vop > 10 else "NO"
        recs = e.recomendaciones(fenotipo, riesgo, vop, pas, pad)

        # ====== Resultado visual ======
        _card_open("🎯 Resultado del análisis")
        st.markdown(
            f"""
            <div class="vh-result-pill" style="background:{color}15; color:{color};
                                               border:1px solid {color}55;">
                <span style="font-size:22px;">●</span>
                <span>Fenotipo: <strong>{fenotipo}</strong></span>
            </div>
            """,
            unsafe_allow_html=True,
        )
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("VOP medida", f"{vop} m/s", delta=f"vs p50: {round(vop-p50,2)}")
        c2.metric("Percentil 50", f"{p50} m/s")
        c3.metric("Fenotipo", fenotipo.split(" (")[0])
        c4.metric("Riesgo CV", riesgo)
        c5, c6, c7, c8 = st.columns(4)
        c5.metric("Presión de Pulso", f"{pp} mmHg")
        c6.metric("PAM", f"{pam} mmHg")
        c7.metric("Edad vascular", f"{edad_vasc} a", delta=f"{edad_vasc-edad:+d} a")
        c8.metric("LOB (VOP>10)", lob)
        _card_close()

        # ====== Gráfico interactivo Plotly ======
        if _PLOTLY_OK:
            _card_open("📈 Lámina interactiva (Plotly)")
            fig_plotly = construir_grafico_plotly(
                edad, sexo, vop, p10, p25, p50, p75, p90,
                color, edad_vasc, pas, pad, riesgo
            )
            if fig_plotly is not None:
                st.plotly_chart(fig_plotly, use_container_width=True,
                                config={"displaylogo": False,
                                        "modeBarButtonsToRemove": ["lasso2d", "select2d"]})
            _card_close()

        try:
            # ====== Lámina matplotlib editorial (también va al PDF) ======
            chart_buf = construir_grafico_didactico(
                edad, sexo, vop, p10, p25, p50, p75, p90,
                color, edad_vasc, pas=pas, pad=pad, riesgo=riesgo)

            with st.expander("🖼️ Lámina didáctica para el PDF (previa)"):
                try:
                    st.image(chart_buf, use_container_width=True)
                except TypeError:
                    chart_buf.seek(0)
                    st.image(chart_buf, use_column_width=True)
            chart_buf.seek(0)

            datos = {"nombre": nombre or "Sin nombre",
                     "documento": documento,
                     "edad": edad, "sexo": sexo,
                     "medico_solicitante": medico_solicitante}
            res = {"vop": vop, "pas": pas, "pad": pad, "pp": pp, "pam": pam,
                   "p10": p10, "p50": p50, "p90": p90, "lob": lob,
                   "edad": edad, "edad_vasc": edad_vasc, "riesgo": riesgo}
            perfil_pdf = _perfil_usuario_actual(st.session_state.get("username", ""))
            pdf_bytes = construir_pdf(datos, res, fenotipo, riesgo, recs,
                                      chart_buf, profesional,
                                      ram.get("_curva_cf_png"),
                                      firma_path=perfil_pdf.get("firma_path"),
                                      sello_path=perfil_pdf.get("sello_path"))

            _card_open("📄 Informe PDF profesional")
            col_a, col_b = st.columns([2, 1])
            with col_a:
                st.markdown(
                    "El informe integra 3 hojas A4: resumen clínico + evidencia CF "
                    "original, lámina didáctica + interpretación, recomendaciones + "
                    "cierre + firma y sello digital del usuario."
                )
            with col_b:
                st.download_button(
                    "⬇ Descargar Informe PDF",
                    data=pdf_bytes,
                    file_name=f"Informe_Vascular_{(nombre or 'paciente').replace(' ', '_')}_"
                              f"{datetime.date.today()}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            _card_close()

            # ====== Recomendaciones ======
            _card_open("📝 Recomendaciones clínicas")
            for i, r in enumerate(recs, 1):
                st.markdown(f"**{i}.** {r}")
            _card_close()

        except Exception as exc:
            st.error(f"Error generando el informe: {exc}")
            with st.expander("Detalle técnico"):
                st.code(traceback.format_exc())
            st.stop()

        # ====== Persistencia ======
        usuario_sesion = st.session_state.get("username", "")
        rol_sesion = st.session_state.get("user_role", "medico")
        fuente_vop = "VOP cf medida" if vop_medida > 0 else "VOP cf recalculada desde distancia/tiempo"
        registro = {
            "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
            "fecha_estudio": str(datetime.date.today()),
            "usuario": usuario_sesion,
            "usuario_id": _usuario_seguro(usuario_sesion),
            "rol": rol_sesion,
            "paciente": nombre,
            "documento": documento,
            "edad": edad,
            "sexo": sexo,
            "medico_solicitante": medico_solicitante,
            "vop_cf_ms": vop,
            "distancia_cf_cm": distancia,
            "tiempo_transito_cf_ms": tiempo,
            "pas_mmhg": pas,
            "pad_mmhg": pad,
            "pam_mmhg": pam,
            "pp_mmhg": pp,
            "p10_ms": p10,
            "p25_ms": p25,
            "p50_ms": p50,
            "p75_ms": p75,
            "p90_ms": p90,
            "edad_vascular": edad_vasc,
            "fenotipo_vascular_unico": fenotipo,
            "riesgo_cv": riesgo,
            "lob_vop_mayor_10": "SI" if vop > 10 else "NO",
            "riesgo_elevado_por_vop_mayor_10": "SI" if vop > 10 else "NO",
            "fuente_vop": fuente_vop,
            "archivo_importado": st.session_state.get("archivo_importado_nombre", ""),
        }
        try:
            _guardar_registro_paciente(registro)
            st.toast("Registro guardado en la base histórica del usuario", icon="✅")
        except Exception as exc:
            st.warning(f"El informe se generó, pero no se pudo guardar el registro histórico: {exc}")


def _pantalla_historial():
    rol = st.session_state.get("user_role", "medico")
    usuario = st.session_state.get("username", "")
    if rol == "admin":
        _hero("Gestión de Datos — Administrador",
              "Vista global: incluye pacientes registrados por todos los usuarios.")
    else:
        _hero("Mis pacientes — Exportación Excel",
              "Vista individual: incluye solo los pacientes cargados por el usuario actual.")

    df = _df_registros(usuario, rol)
    if not df.empty:
        # KPIs rápidos
        col_a, col_b, col_c, col_d = st.columns(4)
        col_a.metric("Pacientes", len(df))
        if "vop_cf_ms" in df.columns:
            try:
                col_b.metric("VOP cf promedio",
                             f"{round(pd.to_numeric(df['vop_cf_ms'], errors='coerce').mean(), 2)} m/s")
            except Exception:
                col_b.metric("VOP cf promedio", "—")
        if "fenotipo_vascular_unico" in df.columns:
            top = df["fenotipo_vascular_unico"].mode()
            col_c.metric("Fenotipo + frecuente",
                         (top.iloc[0].split(" (")[0] if len(top) else "—"))
        if "lob_vop_mayor_10" in df.columns:
            col_d.metric("Con LOB",
                         int((df["lob_vop_mayor_10"] == "SI").sum()))

        _card_open("📊 Registros")
        try:
            st.dataframe(df, use_container_width=True)
        except TypeError:
            st.dataframe(df)
        _card_close()

        excel_data = _excel_bytes_registros(df, "Pacientes_VOP")
        if rol == "admin":
            label = "📥 Excel: exportar pacientes de TODOS los usuarios"
            fname = f"Base_VOP_Todos_los_Usuarios_{datetime.date.today()}.xlsx"
        else:
            label = "📥 Excel: exportar mis pacientes"
            fname = f"Base_VOP_{_usuario_seguro(usuario)}_{datetime.date.today()}.xlsx"
        st.download_button(
            label=label,
            data=excel_data,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        if rol == "admin":
            _card_open("👥 Resumen por usuario")
            resumen = (
                df.groupby("usuario", dropna=False)
                .agg(pacientes=("paciente", "count"), vop_promedio=("vop_cf_ms", "mean"))
                .reset_index()
            )
            try:
                st.dataframe(resumen, use_container_width=True)
            except TypeError:
                st.dataframe(resumen)
            st.download_button(
                label="📥 Excel: exportar resumen por usuario",
                data=_excel_bytes_registros(resumen, "Resumen_Usuarios"),
                file_name=f"Resumen_VOP_Usuarios_{datetime.date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            _card_close()
    else:
        st.info("No hay registros disponibles para exportar.")


def main():
    _inject_premium_css()

    if "auth" not in st.session_state:
        st.session_state.auth = False

    if not st.session_state.auth:
        _pantalla_login()
        return

    usuario_actual = st.session_state.get("username", "")
    profesional, choice = _sidebar(usuario_actual)

    # El selectbox del sidebar trae el menú con emojis; normalizamos:
    if "Administrar" in choice:
        _pantalla_admin()
    elif "Historial" in choice:
        _pantalla_historial()
    else:
        _pantalla_nuevo_estudio(profesional)


if __name__ == "__main__":
    main()


